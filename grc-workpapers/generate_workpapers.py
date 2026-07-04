"""
Big 4 style GRC / Oracle Risk Management Cloud (ITGC + ITAC) Audit Workpaper Package.
Generates a multi-tab Excel workbook with sample data, workings and supporting evidence
references covering P2P, O2C, R2R, Inventory & Payroll cycles, an 87-rule SoD ruleset,
ATC transaction monitoring models and reusable ITGC test templates.
For educational / template use - replace sample data with actual engagement data.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------- Palette & styling ----------
NAVY = "1F3864"; BLUE = "2E5496"; LIGHT = "D9E1F2"; GREY = "F2F2F2"
GREEN = "C6EFCE"; RED = "FFC7CE"; AMBER = "FFEB9C"; WHITE = "FFFFFF"; CRIT = "FF0000"

thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def title_cell(ws, cell, text, size=14):
    ws[cell] = text
    ws[cell].font = Font(bold=True, size=size, color=NAVY)

def header_row(ws, row, headers, start_col=1, fill=BLUE, fontcolor=WHITE):
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=start_col + i, value=h)
        c.font = Font(bold=True, color=fontcolor, size=10)
        c.fill = PatternFill("solid", fgColor=fill)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border

def data_rows(ws, start_row, rows, start_col=1, zebra=True, colormap=None):
    for r, row in enumerate(rows):
        for i, val in enumerate(row):
            c = ws.cell(row=start_row + r, column=start_col + i, value=val)
            c.border = border
            c.alignment = Alignment(vertical="center", wrap_text=True)
            c.font = Font(size=10)
            if zebra and r % 2 == 1:
                c.fill = PatternFill("solid", fgColor=GREY)
            if colormap and i in colormap and str(val) in colormap[i]:
                c.fill = PatternFill("solid", fgColor=colormap[i][str(val)])
    return start_row + len(rows)

def set_widths(ws, widths, start_col=1):
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(start_col + i)].width = w

def note(ws, cell, text, italic=True):
    ws[cell] = text
    ws[cell].font = Font(italic=italic, size=9, color="595959")
    ws[cell].alignment = Alignment(wrap_text=True, vertical="top")

# Reusable RCM builder (13 columns, same schema across cycles)
RCM_HEADERS = ["Risk ID", "Sub-process", "Risk Description", "Rating", "Control ID",
               "Control Description", "Type", "Nature", "Frequency", "Owner",
               "SoD Impact", "Assertion", "Test Procedure"]
RCM_WIDTHS = [4, 10, 12, 30, 8, 12, 34, 12, 12, 10, 14, 9, 15, 30]
RCM_CMAP = {3: {"High": RED, "Med": AMBER, "Low": LIGHT}, 8: {"Automated": GREEN}, 10: {"Yes": AMBER}}

def build_rcm(wb, tab, tabcolor, title, rows):
    ws = wb.create_sheet(tab)
    ws.sheet_properties.tabColor = tabcolor
    title_cell(ws, "B2", title)
    header_row(ws, 4, RCM_HEADERS, start_col=2)
    data_rows(ws, 5, rows, start_col=2, colormap=RCM_CMAP)
    set_widths(ws, RCM_WIDTHS, start_col=1)
    ws.freeze_panes = "A5"
    return ws


wb = Workbook()

# =========================================================
# TAB: INDEX
# =========================================================
ws = wb.active
ws.title = "00_Index"
ws.sheet_properties.tabColor = NAVY
title_cell(ws, "B2", "GRC / ITGC-ITAC AUDIT WORKPAPER PACKAGE", 16)
note(ws, "B3", "Client: ABC Manufacturing Ltd.  |  Period: FY 2025-26  |  System: Oracle Fusion Cloud ERP (24C)")
note(ws, "B4", "Prepared for education/template use. Replace sample data with actual engagement data.")
header_row(ws, 6, ["WP Ref", "Section", "Tab", "Purpose"], start_col=2)
idx = [
    ["A-1", "Planning", "01_Scoping_Memo", "Scope, in-scope apps & processes, materiality"],
    ["A-2", "Planning", "02_IPE_Register", "Completeness & accuracy of reports used"],
    ["RCM", "Control Design", "03_RCM_P2P", "Risk Control Matrix - Procure to Pay"],
    ["RCM", "Control Design", "04_RCM_O2C", "Risk Control Matrix - Order to Cash"],
    ["RCM", "Control Design", "05_RCM_R2R", "Risk Control Matrix - Record to Report"],
    ["RCM", "Control Design", "06_RCM_Inventory", "Risk Control Matrix - Inventory"],
    ["RCM", "Control Design", "07_RCM_Payroll", "Risk Control Matrix - Payroll (Hire to Retire)"],
    ["C", "ITGC", "08_ITGC_Testing", "Worked examples - Access & Change testing"],
    ["C", "ITGC", "09_ITGC_Templates", "Reusable test templates - 4 ITGC domains"],
    ["D", "ITAC", "10_ITAC_Testing", "Automated application controls (3-way match etc.)"],
    ["D", "ATC", "11_ATC_Transaction_Models", "Transaction monitoring models (fraud detection)"],
    ["E-1", "SoD", "12_SoD_Ruleset", "87-rule production Segregation of Duties matrix"],
    ["E-2", "SoD", "13_SoD_Conflict_Analysis", "AAC scan results & disposition"],
    ["E-3", "Security", "14_Role_Design", "Custom role design worksheet"],
    ["F", "Substantive", "15_Test_of_Details", "Reperformance / recalculation"],
    ["G", "Deficiency", "16_Deficiency_Log", "Exception log & severity assessment"],
    ["H", "Remediation", "17_Remediation_Roadmap", "Management action plan & phasing"],
    ["I", "Reporting", "18_Findings_Report", "Findings for audit committee"],
    ["I", "Monitoring", "19_Monitoring_Dashboard", "Continuous monitoring KPIs & metrics"],
    ["-", "Reference", "20_Sampling_Guidance", "Sample size reference & tickmark legend"],
]
data_rows(ws, 7, idx, start_col=2)
set_widths(ws, [10, 16, 30, 55], start_col=2)
note(ws, "B29", "Companion Word deliverables: Scoping_Memorandum.docx, Findings_Report.docx, AAC_Access_Model_Config_Guide.docx (see grc-workpapers folder).")

# =========================================================
# TAB: SCOPING MEMO
# =========================================================
ws = wb.create_sheet("01_Scoping_Memo")
ws.sheet_properties.tabColor = BLUE
title_cell(ws, "B2", "WP A-1: SCOPING MEMORANDUM")
meta = [
    ["Client", "ABC Manufacturing Ltd."],
    ["Period", "01-Apr-2025 to 31-Mar-2026 (FY 2025-26)"],
    ["System", "Oracle Fusion Cloud ERP (Release 24C)"],
    ["Objective", "Evaluate design & operating effectiveness of ITGC & ITAC over financially significant applications supporting ICFR / SOX 404"],
    ["Testing Approach", "Controls reliance + limited substantive"],
    ["Overall Materiality", "INR 12.5 Cr"],
    ["Performance Materiality", "INR 8.0 Cr"],
    ["Prepared By", "________________   Date: __/__/____"],
    ["Reviewed By", "________________   Date: __/__/____"],
]
r = 4
for k, v in meta:
    ws.cell(row=r, column=2, value=k).font = Font(bold=True, size=10)
    c = ws.cell(row=r, column=3, value=v); c.alignment = Alignment(wrap_text=True, vertical="center"); c.font = Font(size=10)
    r += 1
set_widths(ws, [4, 24, 70], start_col=1)
ws.cell(row=r+1, column=2, value="IN-SCOPE APPLICATIONS").font = Font(bold=True, color=NAVY, size=11)
header_row(ws, r+2, ["Application / Module", "Significant?", "Basis"], start_col=2)
apps = [
    ["Oracle General Ledger", "Yes", "Financial statement close"],
    ["Oracle Payables", "Yes", "Expenses / Accounts Payable"],
    ["Oracle Receivables", "Yes", "Revenue / Accounts Receivable"],
    ["Oracle Fixed Assets", "Yes", "Depreciation / PP&E"],
    ["Oracle Inventory", "Yes", "Cost of Goods Sold"],
    ["Oracle HCM Payroll", "Yes", "Payroll expense"],
]
data_rows(ws, r+3, apps, start_col=2, colormap={1: {"Yes": GREEN}})
set_widths(ws, [4, 24, 14, 40], start_col=1)

# =========================================================
# TAB: IPE REGISTER
# =========================================================
ws = wb.create_sheet("02_IPE_Register")
ws.sheet_properties.tabColor = BLUE
title_cell(ws, "B2", "WP A-2: IPE (Information Produced by Entity) REGISTER")
note(ws, "B3", "Every report used in testing must have Completeness & Accuracy (C&A) validated - most common Big 4 review comment.")
header_row(ws, 5, ["Report Name", "Source System", "Used For", "Row Count", "C&A Validated?", "Validation Method", "Evidence Ref"], start_col=2)
ipe = [
    ["AP Aging Report", "Oracle Payables", "Sample selection - AP", "1,842", "Yes", "Reconciled to GL; parameters screenshot", "A-2.1"],
    ["Change Log Extract", "Oracle FSM", "ITGC change testing", "18", "Yes", "Full population; sequence completeness check", "A-2.2"],
    ["User Access Listing", "Security Console", "SoD / access testing", "342", "Yes", "Reconciled to HR active employee list", "A-2.3"],
    ["Fixed Asset Register", "Oracle Fixed Assets", "Depreciation reperformance", "1,240", "Yes", "Reconciled to GL net book value", "A-2.4"],
    ["Invoice Register", "Oracle Payables", "3-way match testing", "9,655", "Yes", "Reconciled to AP subledger control a/c", "A-2.5"],
    ["Payroll Register", "Oracle HCM", "Payroll testing", "612", "Yes", "Reconciled to payroll GL expense", "A-2.6"],
]
data_rows(ws, 6, ipe, start_col=2, colormap={4: {"Yes": GREEN}})
set_widths(ws, [4, 22, 18, 22, 12, 15, 34, 12], start_col=1)

# =========================================================
# RCM TABS
# =========================================================
p2p = [
    ["P2P-R-01","Requisitioning","Purchases without budget/authorization","High","P2P-C-01","System enforces budget check & approval hierarchy on requisition","Preventive","Automated","Per txn","Procurement Mgr","No","Occurrence","Config review + positive/negative test"],
    ["P2P-R-02","Purchase Order","Unauthorized buyer creates PO","High","P2P-C-02","Only authorized buyers create PO; approval limits enforced","Preventive","Automated","Per txn","Procurement Mgr","Yes","Occurrence","Config + access review"],
    ["P2P-R-03","Goods Receipt","Goods recorded not actually received","Med","P2P-C-03","Receiver segregated from buyer; GRN matched to PO","Preventive","IT-Dependent","Per txn","Warehouse Head","Yes","Existence","Walkthrough + sample 25"],
    ["P2P-R-04","Invoicing","Duplicate invoice / payment","High","P2P-C-04","System blocks duplicate invoice number per supplier","Preventive","Automated","Per txn","AP Manager","No","Accuracy","Config + negative test"],
    ["P2P-R-05","Matching","Invoice paid despite PO/GRN mismatch","High","P2P-C-05","Automated 3-way match; hold applied if variance > tolerance","Preventive","Automated","Per txn","AP Manager","No","Accuracy/Valuation","Config + positive/negative test"],
    ["P2P-R-06","Payment","Unauthorized / fraudulent payment","High","P2P-C-06","Payment approval segregated from invoice entry; dual bank control","Preventive","Manual","Per txn","Treasury Head","Yes","Occurrence","Sample 25 + SoD scan"],
    ["P2P-R-07","Master Data","Fake supplier created for fraud","High","P2P-C-07","Supplier creation segregated from payment; new-supplier approval","Preventive","IT-Dependent","Per event","Master Data Team","Yes","Occurrence","Sample + SoD scan"],
    ["P2P-R-08","Reconciliation","AP errors not detected timely","Med","P2P-C-08","Monthly AP subledger to GL reconciliation by independent person","Detective","Manual","Monthly","AP Manager","Yes","Completeness","Sample 2 months"],
]
build_rcm(wb, "03_RCM_P2P", "375623", "RISK CONTROL MATRIX - PROCURE TO PAY (P2P)", p2p)

o2c = [
    ["O2C-R-01","Credit Mgmt","Sales to customers exceeding credit limit","High","O2C-C-01","System enforces credit check & holds order if limit exceeded","Preventive","Automated","Per txn","Credit Manager","Yes","Valuation","Config + negative test"],
    ["O2C-R-02","Order Entry","Unauthorized / invalid sales order","Med","O2C-C-02","Order requires approval per pricing & discount matrix","Preventive","Automated","Per txn","Sales Manager","Yes","Occurrence","Config + sample"],
    ["O2C-R-03","Pricing","Unauthorized discount granted","High","O2C-C-03","Discounts above threshold require approval","Preventive","Automated","Per txn","Sales Manager","No","Accuracy","Config + negative test"],
    ["O2C-R-04","Shipping","Goods shipped but not billed","High","O2C-C-04","System matches shipment to invoice; unbilled-shipment report","Detective","IT-Dependent","Daily","Logistics Head","No","Completeness","Sample + report review"],
    ["O2C-R-05","Invoicing","Invoice inaccurate vs order/shipment","High","O2C-C-05","Auto invoice generation from shipment (AutoInvoice)","Preventive","Automated","Per txn","Billing Lead","No","Accuracy","Config + reperformance"],
    ["O2C-R-06","Cash Application","Receipts misapplied / lapping","Med","O2C-C-06","Auto-match of receipts to invoices; exceptions reviewed","Detective","IT-Dependent","Daily","AR Manager","Yes","Accuracy","Sample 25"],
    ["O2C-R-07","Credit Memo","Unauthorized credit / write-off","High","O2C-C-07","Credit memo & write-off require approval hierarchy","Preventive","Automated","Per txn","AR Manager","Yes","Occurrence","Config + sample"],
    ["O2C-R-08","Reconciliation","AR errors not detected timely","Med","O2C-C-08","Monthly AR subledger to GL reconciliation, independent","Detective","Manual","Monthly","AR Manager","Yes","Completeness","Sample 2 months"],
]
build_rcm(wb, "04_RCM_O2C", "1F6E43", "RISK CONTROL MATRIX - ORDER TO CASH (O2C)", o2c)

r2r = [
    ["R2R-R-01","Journal Entry","Unauthorized / fraudulent journal","High","R2R-C-01","All manual journals require independent approval","Preventive","Automated","Per txn","Controller","Yes","Occurrence","Config + sample 25"],
    ["R2R-R-02","Period Close","Entries posted to closed period","High","R2R-C-02","System locks closed GL periods; open/close restricted","Preventive","Automated","Per period","Controller","Yes","Cut-off","Config + access review"],
    ["R2R-R-03","Reconciliation","Balance sheet accounts not reconciled","High","R2R-C-03","Monthly recon prepared & independently reviewed","Detective","Manual","Monthly","GL Manager","Yes","Existence/Valuation","Sample 2 months"],
    ["R2R-R-04","Chart of Accounts","Improper COA changes","Med","R2R-C-04","COA changes require approval & change control","Preventive","IT-Dependent","Per event","Controller","Yes","Classification","Sample"],
    ["R2R-R-05","Intercompany","Unbalanced intercompany entries","Med","R2R-C-05","IC transactions auto-balanced & reconciled","Detective","Automated","Monthly","GL Manager","No","Completeness","Report review"],
    ["R2R-R-06","Accruals","Inaccurate accruals / estimates","Med","R2R-C-06","Accruals reviewed & approved with support","Preventive","Manual","Monthly","Controller","No","Accuracy","Sample"],
    ["R2R-R-07","FX Rates","Incorrect currency rates applied","Med","R2R-C-07","Rates auto-loaded from approved source; exceptions flagged","Preventive","Automated","Daily","Treasury","No","Valuation","Config + sample"],
    ["R2R-R-08","Reporting","Misstated financial reports","High","R2R-C-08","Reports reconciled to GL; reviewed before issuance","Detective","Manual","Quarterly","CFO","No","Presentation","Sample"],
]
build_rcm(wb, "05_RCM_R2R", "385723", "RISK CONTROL MATRIX - RECORD TO REPORT (R2R)", r2r)

inv = [
    ["INV-R-01","Receiving","Inventory recorded but not received","Med","INV-C-01","Receipt matched to PO; receiver segregated","Preventive","IT-Dependent","Per txn","Warehouse Head","Yes","Existence","Sample 25"],
    ["INV-R-02","Issue","Unauthorized inventory issue","Med","INV-C-02","Issues require authorized requisition","Preventive","Automated","Per txn","Store Manager","Yes","Occurrence","Sample"],
    ["INV-R-03","Adjustments","Fraudulent inventory adjustment","High","INV-C-03","Adjustments above threshold require approval","Preventive","Automated","Per txn","Inventory Manager","Yes","Existence","Config + sample"],
    ["INV-R-04","Cycle Count","Discrepancies not detected","Med","INV-C-04","Periodic cycle counts; variances investigated & approved","Detective","Manual","Monthly","Inventory Manager","Yes","Existence","Sample counts"],
    ["INV-R-05","Valuation","Inventory misvalued","High","INV-C-05","System applies standard/actual cost; cost changes approved","Preventive","Automated","Per event","Cost Accountant","No","Valuation","Config + reperformance"],
    ["INV-R-06","Obsolescence","Obsolete stock not provided for","Med","INV-C-06","Aging report reviewed; provision approved","Detective","Manual","Quarterly","Cost Accountant","No","Valuation","Report review"],
    ["INV-R-07","Transfer","Stock diverted during transfer","Med","INV-C-07","Inter-org transfers require approval & in-transit tracking","Preventive","IT-Dependent","Per txn","Logistics Head","Yes","Existence","Sample"],
    ["INV-R-08","Reconciliation","Inventory subledger vs GL mismatch","Med","INV-C-08","Monthly inventory to GL reconciliation","Detective","Manual","Monthly","Cost Accountant","Yes","Completeness","Sample 2 months"],
]
build_rcm(wb, "06_RCM_Inventory", "806000", "RISK CONTROL MATRIX - INVENTORY", inv)

pay = [
    ["PAY-R-01","New Hire","Ghost employee added to payroll","High","PAY-C-01","New-hire setup segregated from payroll run; HR approval","Preventive","IT-Dependent","Per event","HR Manager","Yes","Occurrence","Sample"],
    ["PAY-R-02","Master Data","Unauthorized salary change","High","PAY-C-02","Salary changes require approval & audit trail","Preventive","IT-Dependent","Per event","HR Manager","Yes","Accuracy","Sample 25"],
    ["PAY-R-03","Time & Attendance","Inflated hours paid","Med","PAY-C-03","Timesheets approved by supervisor before payroll","Preventive","Manual","Per period","Line Managers","Yes","Accuracy","Sample"],
    ["PAY-R-04","Payroll Run","Unauthorized payroll processed","High","PAY-C-04","Payroll run reviewed & approved before disbursement","Preventive","Manual","Per cycle","Payroll Manager","Yes","Occurrence","Sample all runs"],
    ["PAY-R-05","Disbursement","Salary diverted to wrong account","High","PAY-C-05","Bank-detail changes segregated from disbursement; dual control","Preventive","IT-Dependent","Per event","Payroll/Treasury","Yes","Occurrence","Sample + SoD"],
    ["PAY-R-06","Termination","Terminated employee still paid","Med","PAY-C-06","Termination triggers timely payroll deactivation","Preventive","IT-Dependent","Per event","HR Manager","Yes","Occurrence","Sample"],
    ["PAY-R-07","Statutory","Incorrect tax / statutory deduction","Med","PAY-C-07","System auto-calculates statutory deductions per config","Preventive","Automated","Per cycle","Payroll Manager","No","Accuracy","Config + reperformance"],
    ["PAY-R-08","Reconciliation","Payroll not reconciled to GL","Med","PAY-C-08","Payroll register reconciled to GL each cycle","Detective","Manual","Per cycle","Payroll Manager","Yes","Completeness","Sample 2 cycles"],
]
build_rcm(wb, "07_RCM_Payroll", "9C5700", "RISK CONTROL MATRIX - PAYROLL (HIRE TO RETIRE)", pay)

# =========================================================
# TAB: ITGC TESTING (worked examples)
# =========================================================
ws = wb.create_sheet("08_ITGC_Testing")
ws.sheet_properties.tabColor = "7030A0"
title_cell(ws, "B2", "WP C: ITGC TESTING - Access to Programs & Data (Worked Example)")
note(ws, "B3", "ITGC-01: New user access authorized before provisioning. Population=47, Sample=25 (random). Y=verified, N=exception.")
header_row(ws, 5, ["Sr", "User ID", "Access Request Form?", "Approved By", "Approval before Grant?", "Result", "Evidence Ref"], start_col=2)
itgc = [
    [1,"jsmith","Y","Reporting Manager","Y","Pass","C-1.1"],
    [2,"rpatel","Y","Dept Head","Y","Pass","C-1.2"],
    [3,"akhan","N","N/A","N","FAIL","C-1.3"],
    [4,"slodha","Y","Reporting Manager","Y","Pass","C-1.4"],
    [5,"dmehta","Y","Dept Head","Y","Pass","C-1.5"],
    [6,"vnair","Y","Reporting Manager","Y","Pass","C-1.6"],
    ["...","...","...","...","...","...","..."],
    [25,"mgupta","Y","Reporting Manager","Y","Pass","C-1.25"],
]
end = data_rows(ws, 6, itgc, start_col=2, colormap={4:{"N":RED,"Y":GREEN},5:{"FAIL":RED,"Pass":GREEN}})
ws.cell(row=end+1, column=2, value="EXCEPTIONS: 1 of 25 (akhan - no approval). Exception rate 4%.").font = Font(bold=True, color="9C0006", size=10)
ws.cell(row=end+2, column=2, value="CONCLUSION: Control NOT operating effectively. Refer Deficiency Log D-03.").font = Font(bold=True, size=10)
set_widths(ws, [4,6,12,20,18,22,12,12], start_col=1)
sr = end + 5
ws.cell(row=sr, column=2, value="WP C-2: CHANGE MANAGEMENT (Worked Example)").font = Font(bold=True, color=NAVY, size=12)
note(ws, f"B{sr+1}", "ITGC-05: Changes tested & approved before prod. Pop=18, Sample=8. A1=CR documented, A2=UAT sign-off, A3=Approved before migration, A4=Dev != Migrator.")
header_row(ws, sr+2, ["Sr","CR#","A1","A2","A3","A4","Result"], start_col=2)
chg = [
    [1,"CR-1021","Y","Y","Y","Y","Pass"],[2,"CR-1044","Y","Y","Y","N","FAIL"],
    [3,"CR-1052","Y","Y","Y","Y","Pass"],[4,"CR-1061","Y","Y","Y","Y","Pass"],
    ["...","...","...","...","...","...","..."],[8,"CR-1099","Y","Y","Y","Y","Pass"],
]
end2 = data_rows(ws, sr+3, chg, start_col=2, colormap={5:{"N":RED,"Y":GREEN},6:{"FAIL":RED,"Pass":GREEN}})
ws.cell(row=end2+1, column=2, value="EXCEPTION: CR-1044 - Developer 'dev01' also had migration access (SoD - links to D-05).").font = Font(bold=True, color="9C0006", size=10)

# =========================================================
# TAB: ITGC TEMPLATES (reusable, 4 domains)
# =========================================================
ws = wb.create_sheet("09_ITGC_Templates")
ws.sheet_properties.tabColor = "7030A0"
title_cell(ws, "B2", "WP C: ITGC TEST TEMPLATES (Reusable - 4 Domains)")
note(ws, "B3", "Reusable test-step templates. Copy a row into a testing tab and populate results for the engagement.")
header_row(ws, 5, ["Domain", "Control Ref", "Control Objective", "Control Activity", "Test Steps", "Attributes to Verify", "Frequency", "Suggested Sample"], start_col=2)
tmpl = [
    ["Access Mgmt","AC-01","Access granted only when authorized","New user provisioning requires documented approval","1) Obtain new-user list 2) Select sample 3) Trace to approved access request","Request form exists; approver authorized; approval predates grant","Per event","25"],
    ["Access Mgmt","AC-02","Terminated users promptly disabled","De-provisioning on termination within SLA","1) Obtain leaver list from HR 2) Match to disable date","Account disabled <= SLA days from termination","Per event","25"],
    ["Access Mgmt","AC-03","Access remains appropriate over time","Periodic user access review / certification","1) Obtain certification evidence 2) Verify completeness & sign-off","Review performed; sign-off by owner; exceptions actioned","Quarterly","2-4 cycles"],
    ["Access Mgmt","AC-04","Privileged access is restricted","Admin / super-user access limited & monitored","1) Obtain privileged user list 2) Validate business need","Access justified; approved; activity logged/reviewed","Continuous","All admins"],
    ["Access Mgmt","AC-05","Authentication is enforced","Password policy & MFA configured","1) Obtain config 2) Verify against policy","Min length, complexity, expiry, lockout, MFA enabled","Config","1 (test of one)"],
    ["Change Mgmt","CM-01","Changes are authorized","Change request raised & approved","1) Obtain change log 2) Sample 3) Trace to CR & approval","CR documented; approved before build","Per event","8-25"],
    ["Change Mgmt","CM-02","Changes are tested","UAT / test sign-off before migration","1) Verify test evidence for sample","Test results retained; sign-off obtained","Per event","8-25"],
    ["Change Mgmt","CM-03","Migration is segregated","Developer cannot migrate to production","1) Review migration access 2) Compare to developers","Dev != Migrator; SoD enforced","Per event/Config","All migrators"],
    ["Change Mgmt","CM-04","Emergency changes controlled","Emergency change follows post-approval process","1) Obtain emergency change list 2) Verify post-facto approval","Justified; documented; approved after event","Per event","All emergency"],
    ["IT Operations","OP-01","Batch jobs run reliably","Scheduled jobs monitored; failures resolved","1) Obtain job monitoring log 2) Sample failures","Failures detected; resolved & re-run; escalation","Daily","15-25"],
    ["IT Operations","OP-02","Interfaces process completely","Interface reconciliations / error handling","1) Obtain interface control report 2) Verify recon","Records in = records out; errors resolved","Daily","15-25"],
    ["Backup & Recovery","BK-01","Data is recoverable","Backups scheduled & restoration tested","1) Verify backup schedule/logs 2) Verify restore test","Backups successful; periodic restore test passed","Daily/Periodic","Sample + 1 restore"],
]
data_rows(ws, 6, tmpl, start_col=2)
set_widths(ws, [4,14,10,26,26,34,32,16,14], start_col=1)
ws.freeze_panes = "A6"

# =========================================================
# TAB: ITAC TESTING
# =========================================================
ws = wb.create_sheet("10_ITAC_Testing")
ws.sheet_properties.tabColor = "C55A11"
title_cell(ws, "B2", "WP D: ITAC TESTING - Automated Application Controls")
ws.cell(row=4, column=2, value="D-1: 3-Way Match Control (ITAC-P2P-05)").font = Font(bold=True, color=NAVY, size=12)
ws.cell(row=6, column=2, value="TEST OF DESIGN - Config (Payables Options > Matching)").font = Font(bold=True, size=10)
header_row(ws, 7, ["Config Parameter","Expected","Actual","Result","Evidence"], start_col=2)
tod = [
    ["Invoice Match Option","Purchase Order","Purchase Order","Pass","D-1.1"],
    ["Quantity Tolerance","0%","0%","Pass","D-1.1"],
    ["Price Tolerance","2% (per policy)","2%","Pass","D-1.1"],
    ["Match Approval Level","3-Way","3-Way","Pass","D-1.1"],
]
end = data_rows(ws, 8, tod, start_col=2, colormap={3:{"Pass":GREEN}})
sr = end + 2
ws.cell(row=sr, column=2, value="TEST OF ONE - Positive & Negative testing").font = Font(bold=True, size=10)
header_row(ws, sr+1, ["Test Type","Invoice #","Scenario","Expected Behaviour","Actual","Result","Evidence"], start_col=2)
toe = [
    ["Positive","INV-4521","Qty & price within tolerance","No hold - allow payment","No hold applied","Pass","D-1.2"],
    ["Negative","INV-6698","Price 8% > PO (exceeds 2% tol)","Price hold applied; payment blocked","Price hold auto-applied; blocked","Pass","D-1.3"],
]
end2 = data_rows(ws, sr+2, toe, start_col=2, colormap={5:{"Pass":GREEN}})
ws.cell(row=end2+2, column=2, value="ITGC DEPENDENCY: Config unchanged during period (per change log). Test of one valid for full period.").font = Font(italic=True, size=9, color="595959")
ws.cell(row=end2+3, column=2, value="CONCLUSION: Control operating effectively. No exceptions.").font = Font(bold=True, color="375623", size=10)
set_widths(ws, [4,16,18,30,26,12,12], start_col=1)

# =========================================================
# TAB: ATC TRANSACTION MODELS
# =========================================================
ws = wb.create_sheet("11_ATC_Transaction_Models")
ws.sheet_properties.tabColor = "BF4E11"
title_cell(ws, "B2", "WP D: ATC TRANSACTION MONITORING MODELS")
note(ws, "B3", "Advanced Transaction Controls - continuous monitoring logic to detect anomalies/fraud across the full population (not sampling).")
header_row(ws, 5, ["Model ID","Model Name","Process","Detection Logic","Risk Addressed","Priority","Disposition / Action"], start_col=2)
atc = [
    ["ATC-01","Duplicate Payments","P2P","Same supplier + amount + invoice date within N days","Double / duplicate payment","High","Investigate & recover"],
    ["ATC-02","Ghost Vendor (bank match)","P2P","Supplier bank account = employee bank account","Fictitious vendor fraud","High","Investigate immediately"],
    ["ATC-03","Vendor Address = Employee Address","P2P","Supplier address matches employee HR address","Related-party / fictitious vendor","High","Investigate"],
    ["ATC-04","Duplicate Invoice Number","P2P","Same invoice number for same supplier","Duplicate booking","High","Review before payment"],
    ["ATC-05","Split Transactions","P2P","Multiple POs/invoices just below approval limit, same vendor/day","Approval-limit circumvention","Med","Review"],
    ["ATC-06","Payments Just Below Threshold","P2P","Payment amount within 5% under approval limit","Threshold gaming","Med","Trend analysis"],
    ["ATC-07","Round-Dollar Amounts","P2P/GL","Large exact round-number transactions","Fabricated amounts","Med","Sample review"],
    ["ATC-08","Weekend / Holiday Postings","R2R","Journals posted on non-working days","Unauthorized off-hours activity","Med","Review"],
    ["ATC-09","Manual Journals Above Threshold","R2R","Manual JE > materiality by single user","Fraud / error risk","High","Review support"],
    ["ATC-10","Dormant Vendor Reactivation","P2P","Inactive vendor reactivated then paid quickly","Dormant vendor fraud","Med","Investigate"],
    ["ATC-11","Duplicate Employees","Payroll","Same bank/PAN/address across employee IDs","Ghost / duplicate employee","High","Investigate"],
    ["ATC-12","Post-Termination Payments","Payroll","Payment to employee after termination date","Ghost payroll","High","Recover"],
    ["ATC-13","New Vendor + Immediate High Payment","P2P","Vendor created and paid large amount within days","Rushed fraudulent payment","High","Investigate"],
    ["ATC-14","Credit Memo Spike","O2C","Unusual volume/value of credit memos by a user","Revenue manipulation / kickback","Med","Review"],
    ["ATC-15","Sequential Invoice Anomaly","O2C","Gaps or duplicates in invoice number sequence","Completeness / fraud","Med","Review"],
]
data_rows(ws, 6, atc, start_col=2, colormap={5:{"High":RED,"Med":AMBER}})
set_widths(ws, [4,10,26,10,44,26,10,24], start_col=1)
ws.freeze_panes = "A6"

# =========================================================
# TAB: SoD RULESET (87 rules)
# =========================================================
ws = wb.create_sheet("12_SoD_Ruleset")
ws.sheet_properties.tabColor = "C00000"
title_cell(ws, "B2", "WP E-1: SEGREGATION OF DUTIES (SoD) RULESET - 87 RULES")
note(ws, "B3", "Production-grade ruleset used to configure Oracle AAC Access Models. Core principle: no single person controls a complete transaction cycle.")
header_row(ws, 5, ["Rule ID","Process","Function A","Function B (Conflicts)","Severity","Business Impact"], start_col=2)
sod = [
    # P2P (1-15)
    ["SOD-01","P2P","Maintain Supplier","Process AP Payment","High","Fake supplier created and self-paid"],
    ["SOD-02","P2P","Enter Invoice","Approve Payment","High","Fictitious invoice self-paid"],
    ["SOD-03","P2P","Enter Purchase Order","Approve Purchase Order","High","Unauthorized purchase self-approved"],
    ["SOD-04","P2P","Maintain Supplier Bank Details","Process Payment","High","Payment diverted to altered bank account"],
    ["SOD-05","P2P","Create Purchase Requisition","Approve Purchase Requisition","Med","Unauthorized spend initiated"],
    ["SOD-06","P2P","Receive Goods (GRN)","Enter Purchase Order","High","Ghost receipt for fictitious PO"],
    ["SOD-07","P2P","Enter Invoice","Maintain Supplier","High","Fictitious supplier plus invoice"],
    ["SOD-08","P2P","Approve Invoice","Process Payment","High","Self-approved invoice paid"],
    ["SOD-09","P2P","Maintain Payment Terms","Process Payment","Med","Favorable terms enabling fraud payment"],
    ["SOD-10","P2P","Release Invoice Hold","Enter Invoice","Med","Self-release of matching holds"],
    ["SOD-11","P2P","Maintain Supplier","Approve Supplier","High","Self-approved new supplier"],
    ["SOD-12","P2P","Enter Manual Payment","Bank Reconciliation","High","Fraudulent payment concealed in recon"],
    ["SOD-13","P2P","Void Payment","Process Payment","Med","Payment manipulation / recycling"],
    ["SOD-14","P2P","Maintain Catalog Price","Approve Purchase Order","Med","Inflated pricing self-approved"],
    ["SOD-15","P2P","Enter AP Credit Memo","Approve AP Credit Memo","Med","Fraudulent supplier credit"],
    # O2C (16-27)
    ["SOD-16","O2C","Maintain Customer","Post Cash Receipt","Med","Misapplied receipts / lapping"],
    ["SOD-17","O2C","Maintain Customer Remit-To","Apply Cash Receipt","Med","Diverted collections"],
    ["SOD-18","O2C","Create Sales Order","Approve Credit Limit","High","Sales beyond approved credit"],
    ["SOD-19","O2C","Enter Sales Invoice","Approve Sales Invoice","Med","Fictitious revenue booked"],
    ["SOD-20","O2C","Create AR Credit Memo","Approve AR Credit Memo","High","Unauthorized write-off / kickback"],
    ["SOD-21","O2C","Post Cash Receipt","Perform AR Reconciliation","Med","Concealed misapplication"],
    ["SOD-22","O2C","Write-off Bad Debt","Approve Write-off","High","Concealed theft of collections"],
    ["SOD-23","O2C","Maintain Pricing / Discount","Approve Sales Order","High","Unauthorized discounts granted"],
    ["SOD-24","O2C","Ship Goods","Create Sales Invoice","Med","Shipment without billing"],
    ["SOD-25","O2C","Maintain Customer Credit Limit","Create Sales Order","High","Self-set credit plus order"],
    ["SOD-26","O2C","Apply Cash Receipt","Adjust Customer Balance","Med","Customer balance manipulation"],
    ["SOD-27","O2C","Process Refund","Approve Refund","High","Fraudulent refund to self"],
    # R2R (28-37)
    ["SOD-28","R2R","Create Journal Entry","Approve Journal Entry","High","Fraudulent entry self-approved"],
    ["SOD-29","R2R","Post Journal","Perform Account Reconciliation","Med","Errors / fraud concealed in recon"],
    ["SOD-30","R2R","Maintain Chart of Accounts","Post Journal","Med","Misclassification / hidden entries"],
    ["SOD-31","R2R","Open/Close GL Period","Post Journal","High","Backdated entries in closed period"],
    ["SOD-32","R2R","Create Recurring Journal","Approve Recurring Journal","Med","Automated fraudulent entries"],
    ["SOD-33","R2R","Maintain Allocation Rules","Post Journal","Med","Manipulated cost allocations"],
    ["SOD-34","R2R","Enter Intercompany Txn","Approve Intercompany Txn","Med","Unbalanced / fraudulent IC entries"],
    ["SOD-35","R2R","Maintain Currency Rates","Post Journal","Med","FX manipulation for gains"],
    ["SOD-36","R2R","Create Manual Journal","Maintain Chart of Accounts","Med","Self-create account plus post"],
    ["SOD-37","R2R","Approve Journal","Perform Reconciliation","Med","Self-review of own approvals"],
    # H2R / Payroll (38-49)
    ["SOD-38","Payroll","Maintain Employee Master","Approve Payroll Run","High","Ghost employee paid"],
    ["SOD-39","Payroll","Maintain Salary / Comp","Approve Payroll","High","Inflated pay self-approved"],
    ["SOD-40","Payroll","Enter Time & Attendance","Approve Timesheet","Med","Inflated hours self-approved"],
    ["SOD-41","Payroll","Add New Hire","Process Payroll Payment","High","Fictitious employee payment"],
    ["SOD-42","Payroll","Maintain Employee Bank Details","Process Payroll Payment","High","Salary diverted"],
    ["SOD-43","Payroll","Process Termination","Maintain Employee Master","Med","Terminated employee kept active"],
    ["SOD-44","Payroll","Maintain Payroll Deductions","Approve Payroll Run","Med","Manipulated deductions"],
    ["SOD-45","Payroll","Approve Payroll","Disburse Payroll","High","Self-approved and disbursed"],
    ["SOD-46","Payroll","Maintain Benefits Enrollment","Approve Benefits","Med","Unauthorized benefits"],
    ["SOD-47","Payroll","Run Payroll","Reconcile Payroll to GL","Med","Concealed payroll errors"],
    ["SOD-48","Payroll","Maintain Tax Setup","Approve Payroll","Med","Statutory manipulation"],
    ["SOD-49","Payroll","Enter Bonus / Incentive","Approve Bonus","High","Self-awarded bonus"],
    # Inventory (50-57)
    ["SOD-50","Inventory","Adjust Inventory Quantity","Physical Custody of Stock","Med","Theft masked by adjustment"],
    ["SOD-51","Inventory","Receive Inventory","Approve Inventory Adjustment","Med","Concealed receipt discrepancy"],
    ["SOD-52","Inventory","Perform Cycle Count","Approve Count Adjustment","Med","Self-approved count variance"],
    ["SOD-53","Inventory","Issue Inventory","Adjust Inventory","Med","Unauthorized issue concealed"],
    ["SOD-54","Inventory","Maintain Item Cost","Approve Inventory Valuation","Med","Manipulated valuation"],
    ["SOD-55","Inventory","Create Inventory Transfer","Approve Transfer","Med","Diverted stock"],
    ["SOD-56","Inventory","Write-off Obsolete Stock","Approve Write-off","Med","Theft via obsolescence write-off"],
    ["SOD-57","Inventory","Maintain Item Master","Adjust Inventory","Med","Fictitious item plus adjustment"],
    # Fixed Assets (58-64)
    ["SOD-58","Fixed Assets","Create / Acquire Asset","Approve Asset Addition","Med","Fictitious asset capitalized"],
    ["SOD-59","Fixed Assets","Retire / Dispose Asset","Approve Asset Disposal","Med","Unauthorized asset disposal"],
    ["SOD-60","Fixed Assets","Maintain Depreciation Method","Run Depreciation","Med","Manipulated depreciation"],
    ["SOD-61","Fixed Assets","Transfer Asset","Approve Asset Transfer","Low","Untracked asset movement"],
    ["SOD-62","Fixed Assets","Maintain Asset Cost","Approve Capitalization","Med","Inflated capitalization"],
    ["SOD-63","Fixed Assets","Perform Physical Verification","Adjust Asset Register","Med","Concealed missing assets"],
    ["SOD-64","Fixed Assets","Create Asset","Physical Custody of Asset","Med","Ghost asset"],
    # Treasury / Cash (65-72)
    ["SOD-65","Treasury","Maintain Bank Account","Make Payment","High","Payment to unauthorized account"],
    ["SOD-66","Treasury","Initiate Wire Transfer","Approve Wire Transfer","High","Unauthorized fund transfer"],
    ["SOD-67","Treasury","Perform Bank Reconciliation","Process Payment","High","Fraud concealed in recon"],
    ["SOD-68","Treasury","Maintain Signatory","Approve Payment","High","Self-added signatory"],
    ["SOD-69","Treasury","Enter Investment Txn","Approve Investment","Med","Unauthorized investment"],
    ["SOD-70","Treasury","Create Payment Batch","Release Payment Batch","High","Self-released batch"],
    ["SOD-71","Treasury","Maintain Cash Position","Approve Funding","Med","Manipulated liquidity data"],
    ["SOD-72","Treasury","Record Bank Transaction","Reconcile Bank","Med","Self-review of entries"],
    # Security / System Admin (73-80)
    ["SOD-73","Security","User Administration","Any Financial Transaction","Critical","Self-grant access and post entries"],
    ["SOD-74","Security","Assign Roles","Approve Role Assignment","High","Self-approved privileged access"],
    ["SOD-75","Security","Maintain SoD Ruleset","Investigate SoD Incidents","High","Rules weakened to hide conflicts"],
    ["SOD-76","Security","Create Custom Role","Assign Role to User","High","Self-designed excessive role"],
    ["SOD-77","Security","Manage Approval Workflow","Approve Transactions","High","Bypass approval routing"],
    ["SOD-78","Security","Database / Basis Admin","Application Transaction","Critical","Backend data manipulation"],
    ["SOD-79","Security","Migrate Config to Production","Develop Config","High","Untested changes to production"],
    ["SOD-80","Security","Reset User Password","User Administration","Med","Account takeover risk"],
    # Cross / Master Data (81-87)
    ["SOD-81","Cross","Maintain Tax Rates","Post Transaction","Med","Tax manipulation"],
    ["SOD-82","Cross","Maintain Approval Limits","Approve Transaction","High","Self-raised approval limits"],
    ["SOD-83","Cross","Maintain GL Account Mapping","Post Journal","Med","Misdirected postings"],
    ["SOD-84","Cross","Create Vendor","Create Customer (same party)","Med","Related-party concealment"],
    ["SOD-85","Cross","Maintain Payment Method","Process Payment","Med","Unauthorized payment channel"],
    ["SOD-86","Cross","Maintain Budget","Approve Budget","Med","Self-approved budget for overspend"],
    ["SOD-87","Cross","Maintain Cost Center","Approve Expense","Med","Misallocated / hidden expenses"],
]
data_rows(ws, 6, sod, start_col=2, colormap={4:{"Critical":CRIT,"High":RED,"Med":AMBER,"Low":LIGHT}})
set_widths(ws, [4,10,12,26,26,12,42], start_col=1)
ws.freeze_panes = "A6"

# =========================================================
# TAB: SoD CONFLICT ANALYSIS
# =========================================================
ws = wb.create_sheet("13_SoD_Conflict_Analysis")
ws.sheet_properties.tabColor = "C00000"
title_cell(ws, "B2", "WP E-2: SoD CONFLICT ANALYSIS (Oracle AAC output)")
note(ws, "B3", "Population: 342 active users scanned via Oracle AAC (path-based). 64 users with 1+ SoD violation.")
header_row(ws, 5, ["Rule","# Users","Severity","Disposition","Comment"], start_col=2)
summ = [
    ["SOD-01",12,"High","Remediate","Remove supplier maintenance access"],
    ["SOD-28",8,"High","Mitigate","Independent monthly journal review"],
    ["SOD-03",23,"High","Remediate","Split PO entry vs approval roles"],
    ["SOD-73",3,"Critical","Remediate","Immediate - remove User Admin from finance"],
    ["SOD-02",11,"High","Remediate","Custom role redesign"],
    ["Others",7,"Med","Accept/Mitigate","Compensating controls documented"],
]
end = data_rows(ws, 6, summ, start_col=2, colormap={3:{"Critical":CRIT,"High":RED,"Med":AMBER},4:{"Remediate":LIGHT}})
set_widths(ws, [4,12,10,12,16,45], start_col=1)
sr = end + 2
ws.cell(row=sr, column=2, value="DETAILED CONFLICT SAMPLE (SOD-01) - path-based").font = Font(bold=True, color=NAVY, size=11)
detail = [
    ["User","bshah"],
    ["Conflicting Roles","Custom_AP_Clerk + Supplier_Maintenance"],
    ["Path A","bshah -> AP_Clerk_Job -> Payables_Payment_Duty -> 'Create Payment'"],
    ["Path B","bshah -> Supplier_Job -> Supplier_Mgmt_Duty -> 'Create Supplier'"],
    ["Risk","Can create fake supplier and self-pay"],
    ["Interim Mitigation","AP Manager reviews new-supplier + payment report weekly (CC-01)"],
    ["Remediation","Remove Supplier_Maintenance role by 31-Aug-2026"],
]
rr = sr + 1
for k, v in detail:
    ws.cell(row=rr, column=2, value=k).font = Font(bold=True, size=10)
    ws.cell(row=rr, column=3, value=v).font = Font(size=10); ws.cell(row=rr, column=3).alignment = Alignment(wrap_text=True)
    rr += 1

# =========================================================
# TAB: ROLE DESIGN
# =========================================================
ws = wb.create_sheet("14_Role_Design")
ws.sheet_properties.tabColor = "203864"
title_cell(ws, "B2", "WP E-3: CUSTOM ROLE DESIGN WORKSHEET")
note(ws, "B3", "Custom role XX_AP_MANAGER_CUSTOM (copied from seeded 'Accounts Payable Manager'). Principle: least privilege + SoD clean.")
header_row(ws, 5, ["Duty Role","Action","Reason","SoD Rule Addressed"], start_col=2)
rd = [
    ["Supplier Profile Management Duty","REMOVE","Conflicts with payment","SOD-01"],
    ["Bank Account Setup Duty","REMOVE","Conflicts with payment","SOD-04 / SOD-65"],
    ["Payables Invoice Approval Duty","RETAIN","Core manager function","-"],
    ["Payables Payment Approval Duty","RETAIN","Core manager function","-"],
    ["AP Inquiry & Reporting Duty","RETAIN","Read-only reporting","-"],
    ["User Account Management Duty","REMOVE","Conflicts with any txn","SOD-73"],
]
end = data_rows(ws, 6, rd, start_col=2, colormap={1:{"REMOVE":RED,"RETAIN":GREEN}})
ws.cell(row=end+1, column=2, value="DATA SECURITY: Business Unit = 'India Operations' only.").font = Font(bold=True, size=10)
ws.cell(row=end+2, column=2, value="SoD RE-SCAN AFTER DESIGN (simulated in AAC before deploy): 0 conflicts - PASS.").font = Font(bold=True, color="375623", size=10)
set_widths(ws, [4,34,12,34,18], start_col=1)

# =========================================================
# TAB: TEST OF DETAILS (with formulas)
# =========================================================
ws = wb.create_sheet("15_Test_of_Details")
ws.sheet_properties.tabColor = "548235"
title_cell(ws, "B2", "WP F-1: TEST OF DETAILS - Depreciation Reperformance")
note(ws, "B3", "Objective: verify accuracy of FA depreciation. Population=1,240; Sample=30. Recalc & Result columns auto-compute.")
header_row(ws, 5, ["Sr","Asset ID","Cost (INR)","Life (yrs)","Method","System Dep","Recalc Dep","Diff","Result"], start_col=2)
assets = [
    [1,"FA-1001",500000,5,"SLM",100000],[2,"FA-1044",240000,4,"SLM",60000],
    [3,"FA-1120",750000,5,"SLM",150000],[4,"FA-1355",360000,3,"SLM",120000],
    [5,"FA-1502",900000,10,"SLM",90000],[6,"FA-1780",480000,4,"SLM",120000],
    [7,"FA-2210",900000,3,"SLM",300300],
]
start = 6
for r, a in enumerate(assets):
    row = start + r
    for i, val in enumerate(a):
        c = ws.cell(row=row, column=2 + i, value=val); c.border = border; c.font = Font(size=10)
        c.alignment = Alignment(vertical="center")
    ws.cell(row=row, column=8).value = f"=ROUND(D{row}/E{row},0)"
    ws.cell(row=row, column=9).value = f"=G{row}-H{row}"
    ws.cell(row=row, column=10).value = f'=IF(ABS(I{row})<1,"Pass","EXCEPTION")'
    for cc in range(8, 11):
        ws.cell(row=row, column=cc).border = border; ws.cell(row=row, column=cc).font = Font(size=10)
endrow = start + len(assets)
ws.cell(row=endrow+1, column=2, value="EXCEPTION: FA-2210 - rounding config diff INR 300 (immaterial).").font = Font(bold=True, color="C55A11", size=10)
ws.cell(row=endrow+2, column=2, value="CONCLUSION: Depreciation calculation accurate. Immaterial difference noted.").font = Font(bold=True, size=10)
set_widths(ws, [4,6,12,12,10,10,12,12,10,12], start_col=1)

# =========================================================
# TAB: DEFICIENCY LOG
# =========================================================
ws = wb.create_sheet("16_Deficiency_Log")
ws.sheet_properties.tabColor = "BF8F00"
title_cell(ws, "B2", "WP G: DEFICIENCY / EXCEPTION LOG & SEVERITY")
header_row(ws, 4, ["Def #","Deficiency Description","Ref WP","Could cause misstatement?","Magnitude","Likelihood","Compensating Control?","Severity","Status"], start_col=2)
defs = [
    ["D-03","1 user provisioned without approval evidence","08_ITGC","Yes (limited)","Low","Remote","Periodic access review","Control Deficiency","Open"],
    ["D-05","Developer had prod migration access (SoD)","08_ITGC","Yes","Moderate","Reasonably possible","None effective","Significant Deficiency","Open"],
    ["D-08","3 users: User Admin + financial txn access","13_SoD","Yes","Material","Reasonably possible","None","MATERIAL WEAKNESS","Open"],
    ["D-11","Rounding diff in depreciation config","15_ToD","Yes (immaterial)","Trivial","Remote","GL recon","Control Deficiency","Open"],
]
data_rows(ws, 5, defs, start_col=2, colormap={7:{"MATERIAL WEAKNESS":CRIT,"Significant Deficiency":RED,"Control Deficiency":AMBER}})
note(ws, "B11", "Severity logic: (1) Could it cause misstatement? (2) Magnitude? (3) Likelihood? (4) Compensating controls? -> classify.")
set_widths(ws, [4,8,32,10,18,12,16,18,20,10], start_col=1)

# =========================================================
# TAB: REMEDIATION ROADMAP
# =========================================================
ws = wb.create_sheet("17_Remediation_Roadmap")
ws.sheet_properties.tabColor = "2E75B6"
title_cell(ws, "B2", "WP H: REMEDIATION ROADMAP / MANAGEMENT ACTION PLAN")
header_row(ws, 4, ["Def #","Remediation Action","Priority","Phase","Owner","Target Date","Validation Method","Status"], start_col=2)
rem = [
    ["D-08","Remove User Admin duty from 3 finance users","P1-Critical","Phase 1 (0-30d)","IT Security Head","15-Jul-2026","AAC re-scan = 0 conflicts","In Progress"],
    ["D-05","Segregate developer & migration access","P2-High","Phase 2 (1-3m)","IT Manager","31-Aug-2026","Change log + SoD re-test","Planned"],
    ["SOD-03","Redesign PO custom roles (entry vs approval)","P2-High","Phase 2 (1-3m)","GRC Lead","31-Aug-2026","Role re-scan + UAT","Planned"],
    ["D-03","Enforce access request workflow before grant","P3-Med","Phase 3 (3-6m)","HR + IT","30-Sep-2026","Re-test sample of new users","Planned"],
    ["Monitoring","Deploy AAC/ATC continuous monitoring","P3-Med","Phase 3 (3-6m)","GRC Lead","31-Oct-2026","Dashboard live + incident SLA","Planned"],
    ["Cert","Implement quarterly access certification","P3-Med","Phase 3 (3-6m)","GRC Lead","31-Oct-2026","First cycle completed","Planned"],
]
data_rows(ws, 5, rem, start_col=2, colormap={3:{"P1-Critical":RED,"P2-High":AMBER}})
sr = 13
ws.cell(row=sr, column=2, value="PHASED APPROACH").font = Font(bold=True, color=NAVY, size=11)
ph = [
    ["Phase 1 (0-30 days)","Critical access removal, disable super-users, obvious SoD fixes - quick wins"],
    ["Phase 2 (1-3 months)","Custom role redesign, AAC model deployment, approval workflow fixes"],
    ["Phase 3 (3-6 months)","Continuous monitoring, periodic access certification, training"],
]
rr = sr + 1
for k, v in ph:
    ws.cell(row=rr, column=2, value=k).font = Font(bold=True, size=10)
    ws.cell(row=rr, column=4, value=v).font = Font(size=10); ws.cell(row=rr, column=4).alignment = Alignment(wrap_text=True)
    rr += 1
set_widths(ws, [4,10,34,12,16,16,14,26,12], start_col=1)
ws.freeze_panes = "A5"

# =========================================================
# TAB: FINDINGS REPORT
# =========================================================
ws = wb.create_sheet("18_Findings_Report")
ws.sheet_properties.tabColor = "808080"
title_cell(ws, "B2", "WP I: FINDINGS REPORT (Audit Committee Format)")
header_row(ws, 4, ["Finding #","Title","Rating","Condition","Criteria","Cause","Effect","Recommendation","Mgmt Response"], start_col=2)
find = [
    ["D-08","Excessive privileged access - SoD violation","Material Weakness",
     "3 finance users hold User Administration + transaction posting in Oracle Fusion",
     "SoD principle - admin segregated from transaction processing (Policy SEC-04 / SOX)",
     "Roles granted during migration; never reviewed",
     "Risk of unauthorized access grant & fraudulent entries; potential material misstatement",
     "Remove User Admin duty; implement quarterly access certification",
     "Agreed. Owner: IT Security Head. Target 15-Jul-2026"],
    ["D-05","Inadequate change segregation","Significant Deficiency",
     "Developer had production migration access",
     "Dev must be segregated from prod migration (change mgmt policy)",
     "No enforced segregation in deployment process",
     "Unauthorized / untested changes could reach production",
     "Segregate roles; enforce migration approval gate",
     "Agreed. Owner: IT Manager. Target 31-Aug-2026"],
]
data_rows(ws, 5, find, start_col=2, colormap={2:{"Material Weakness":CRIT,"Significant Deficiency":RED}})
set_widths(ws, [4,8,22,16,26,26,22,28,26,24], start_col=1)

# =========================================================
# TAB: CONTINUOUS MONITORING DASHBOARD
# =========================================================
ws = wb.create_sheet("19_Monitoring_Dashboard")
ws.sheet_properties.tabColor = "1F6E43"
title_cell(ws, "B2", "CONTINUOUS MONITORING DASHBOARD - AAC / ATC METRICS")
note(ws, "B3", "Near-real-time KPIs from Oracle AAC/ATC continuous monitoring. Population scanned: 342 users, 9,655 transactions. Refresh: with each data sync.")

# --- Section 1: Headline KPIs ---
ws.cell(row=5, column=2, value="1. HEADLINE KPIs").font = Font(bold=True, color=NAVY, size=12)
header_row(ws, 6, ["Metric", "Current", "Prior Period", "Target", "Status"], start_col=2)
kpis = [
    ["Total SoD incidents (open)", 64, 78, "< 40", "Off Track"],
    ["New incidents this period", 9, 15, "Decreasing", "On Track"],
    ["Incidents closed this period", 23, 12, "Increasing", "On Track"],
    ["Critical incidents open", 3, 5, "0", "Off Track"],
    ["High incidents open", 46, 55, "< 20", "Off Track"],
    ["Avg age of open incidents (days)", 37, 44, "< 30", "Watch"],
    ["Mean time to remediate (days)", 28, 35, "< 21", "Watch"],
    ["% incidents with mitigating control", "72%", "60%", "> 90%", "Watch"],
    ["False positive rate", "11%", "24%", "< 10%", "Watch"],
    ["Control coverage (processes w/ active controls)", "83%", "70%", "100%", "Watch"],
    ["Users with conflicts / total users", "64 / 342 (19%)", "78 / 340 (23%)", "< 10%", "Off Track"],
    ["Overdue remediations", 4, 9, "0", "Off Track"],
]
cmap_status = {4: {"Off Track": RED, "Watch": AMBER, "On Track": GREEN}}
end = data_rows(ws, 7, kpis, start_col=2, colormap=cmap_status)
set_widths(ws, [4, 44, 16, 16, 14, 12], start_col=1)

# --- Section 2: Incidents by Severity ---
sr = end + 2
ws.cell(row=sr, column=2, value="2. INCIDENTS BY SEVERITY (status breakdown)").font = Font(bold=True, color=NAVY, size=12)
header_row(ws, sr+1, ["Severity", "Open", "In Investigation", "Remediate", "Accepted", "Total"], start_col=2)
sev = [
    ["Critical", 3, 0, 3, 0, "=SUM(C{r}:F{r})"],
    ["High", 24, 12, 8, 2, "=SUM(C{r}:F{r})"],
    ["Medium", 15, 5, 4, 6, "=SUM(C{r}:F{r})"],
    ["Low", 2, 0, 0, 3, "=SUM(C{r}:F{r})"],
]
srow = sr + 2
for i, row in enumerate(sev):
    rr = srow + i
    for j, val in enumerate(row):
        if isinstance(val, str) and val.startswith("=SUM"):
            val = val.replace("{r}", str(rr))
        c = ws.cell(row=rr, column=2 + j, value=val)
        c.border = border; c.font = Font(size=10); c.alignment = Alignment(vertical="center")
        if i % 2 == 1:
            c.fill = PatternFill("solid", fgColor=GREY)
    ws.cell(row=rr, column=2).fill = PatternFill("solid", fgColor={0: CRIT, 1: RED, 2: AMBER, 3: LIGHT}[i])
end2 = srow + len(sev)

# --- Section 3: Incidents by Process ---
sr3 = end2 + 2
ws.cell(row=sr3, column=2, value="3. INCIDENTS BY PROCESS").font = Font(bold=True, color=NAVY, size=12)
header_row(ws, sr3+1, ["Process", "# Incidents", "% of Total", "Top Rule"], start_col=2)
proc = [
    ["P2P (Procure-to-Pay)", 23, "36%", "SOD-03 (PO entry/approval)"],
    ["R2R (Record-to-Report)", 14, "22%", "SOD-28 (JE create/approve)"],
    ["O2C (Order-to-Cash)", 10, "16%", "SOD-20 (credit memo)"],
    ["Payroll", 8, "12%", "SOD-38 (ghost employee)"],
    ["Security / Admin", 5, "8%", "SOD-73 (user admin)"],
    ["Others", 4, "6%", "Various"],
]
end3 = data_rows(ws, sr3+2, proc, start_col=2)

# --- Section 4: Incident Aging ---
sr4 = end3 + 2
ws.cell(row=sr4, column=2, value="4. OPEN INCIDENT AGING").font = Font(bold=True, color=NAVY, size=12)
header_row(ws, sr4+1, ["Age Bucket", "Count", "SLA Breached?"], start_col=2)
aging = [
    ["0 - 30 days", 34, "No"],
    ["31 - 60 days", 18, "Watch"],
    ["61 - 90 days", 8, "Yes"],
    ["90+ days", 4, "Yes - escalate"],
]
end4 = data_rows(ws, sr4+2, aging, start_col=2, colormap={2: {"Yes": RED, "Yes - escalate": RED, "Watch": AMBER, "No": GREEN}})

# --- Section 5: ATC Transaction Monitoring ---
sr5 = end4 + 2
ws.cell(row=sr5, column=2, value="5. ATC TRANSACTION MONITORING (this period)").font = Font(bold=True, color=NAVY, size=12)
header_row(ws, sr5+1, ["Model", "Txns Flagged", "Confirmed Issues", "Value at Risk (INR)", "Status"], start_col=2)
atcm = [
    ["ATC-01 Duplicate Payments", 12, 2, "8,40,000", "Under Recovery"],
    ["ATC-02 Ghost Vendor (bank match)", 3, 1, "12,00,000", "Investigating"],
    ["ATC-12 Post-Termination Payments", 5, 3, "2,15,000", "Recovered"],
    ["ATC-09 Large Manual Journals", 18, 0, "-", "Reviewed - OK"],
    ["ATC-05 Split Transactions", 7, 2, "1,90,000", "Investigating"],
]
end5 = data_rows(ws, sr5+2, atcm, start_col=2, colormap={5: {"Investigating": AMBER, "Under Recovery": AMBER, "Recovered": GREEN, "Reviewed - OK": GREEN}})
set_widths(ws, [4, 44, 16, 16, 18, 16], start_col=1)

ws.cell(row=end5+2, column=2, value="Note: Metrics illustrate a monthly monitoring cadence. Green=on track, Amber=watch, Red=off track/breach.").font = Font(italic=True, size=9, color="595959")

# =========================================================
# TAB: SAMPLING GUIDANCE
# =========================================================
ws = wb.create_sheet("20_Sampling_Guidance")
ws.sheet_properties.tabColor = "A6A6A6"
title_cell(ws, "B2", "REFERENCE: SAMPLING GUIDANCE & TICKMARK LEGEND")
header_row(ws, 4, ["Control Frequency","Approx Population","Suggested Sample Size"], start_col=2)
samp = [
    ["Annual","1","1"],["Quarterly","4","2"],["Monthly","12","2 - 5"],
    ["Weekly","52","5 - 15"],["Daily","250+","15 - 25"],
    ["Multiple times per day","Many","25 - 40"],
    ["Automated (config)","N/A","1 (test of one) + ITGC reliance"],
]
data_rows(ws, 5, samp, start_col=2)
sr = 14
ws.cell(row=sr, column=2, value="TICKMARK LEGEND").font = Font(bold=True, color=NAVY, size=11)
legend = [["Y / tick","Verified to supporting evidence"],["N / cross","Exception noted"],
          ["Pass","Attribute met, no exception"],["FAIL","Attribute not met - deficiency"],
          ["E","Evidence attached (referenced)"],["N/A","Not applicable"]]
rr = sr + 1
for k, v in legend:
    ws.cell(row=rr, column=2, value=k).font = Font(bold=True, size=10)
    ws.cell(row=rr, column=3, value=v).font = Font(size=10)
    rr += 1
set_widths(ws, [4,22,30,20], start_col=1)

out = "/projects/sandbox/Vivek-Kriplani/grc-workpapers/GRC_ITAC_Audit_Workpaper_Package.xlsx"
wb.save(out)
print("Saved:", out)
print("Tabs (%d):" % len(wb.sheetnames), wb.sheetnames)
