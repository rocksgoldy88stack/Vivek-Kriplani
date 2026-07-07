"""
Digital Audit & Assurance - Big 4 Deliverables & Workpapers Workbook Generator
Author perspective: 34-yr experienced Big 4 Audit & Assurance Partner
Generates a multi-sheet Excel workbook with realistic sample data.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import date

# ---------- Style palette (Big 4 "deep blue / teal" corporate feel) ----------
NAVY      = "1F3864"   # titles
BLUE      = "2E5496"   # section headers
TEAL      = "1F6F6B"   # accent
LIGHTBLUE = "D9E1F2"   # header row fill
LIGHTGREY = "F2F2F2"   # zebra
AMBER     = "FFF2CC"   # notes
GREEN_F   = "C6EFCE"; GREEN_T = "006100"
RED_F     = "FFC7CE";  RED_T   = "9C0006"
AMBER_F   = "FFEB9C";  AMBER_T = "9C6500"
WHITE     = "FFFFFF"

thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def hfont(sz=11, color=WHITE, bold=True):
    return Font(name="Calibri", size=sz, bold=bold, color=color)

def fill(hex_):
    return PatternFill("solid", fgColor=hex_)

def style_header_row(ws, row, ncols, fillhex=BLUE, fontcolor=WHITE, sz=11):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill(fillhex)
        cell.font = hfont(sz, fontcolor, True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER

def title_block(ws, title, subtitle, ncols):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    t = ws.cell(row=1, column=1, value=title)
    t.font = hfont(16, WHITE, True); t.fill = fill(NAVY)
    t.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 30
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    s = ws.cell(row=2, column=1, value=subtitle)
    s.font = Font(name="Calibri", size=10, italic=True, color=NAVY)
    s.fill = fill(LIGHTBLUE)
    s.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[2].height = 18

def put_table(ws, start_row, headers, rows, widths=None, zebra=True, header_fill=BLUE):
    style_header_row(ws, start_row, len(headers), header_fill)
    for j, h in enumerate(headers, start=1):
        ws.cell(row=start_row, column=j, value=h)
    r = start_row + 1
    for i, rowdata in enumerate(rows):
        for j, val in enumerate(rowdata, start=1):
            c = ws.cell(row=r, column=j, value=val)
            c.border = BORDER
            c.alignment = Alignment(vertical="top", wrap_text=True)
            if zebra and i % 2 == 1:
                c.fill = fill(LIGHTGREY)
        r += 1
    if widths:
        for j, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(j)].width = w
    return r  # next free row

def add_filter_freeze(ws, header_row, ncols):
    last = get_column_letter(ncols)
    ws.auto_filter.ref = f"A{header_row}:{last}{ws.max_row}"
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

wb = Workbook()

# =========================================================================
# 00 - INDEX / CONTROL SHEET
# =========================================================================
ws = wb.active
ws.title = "00_Index"
title_block(ws, "DIGITAL AUDIT & ASSURANCE  —  ENGAGEMENT FILE",
            "Illustrative Big 4 workpaper & deliverables pack  |  Client: Nimbus Retail Ltd (illustrative)  |  FY ended 31-Mar-2026",
            6)
meta = [
    ["Engagement", "Statutory audit + IT audit (integrated)"],
    ["Client (illustrative)", "Nimbus Retail Ltd — listed, omni-channel retailer"],
    ["Reporting framework", "Ind AS / IFRS; audit under SA (ISA)"],
    ["Engagement partner", "CA (Partner) — 34 yrs experience"],
    ["Group turnover (illustrative)", "INR 4,820 crore"],
    ["Overall materiality (illustrative)", "INR 24.1 crore (0.5% of revenue)"],
    ["File status", "In progress — fieldwork"],
    ["Prepared / Reviewed", "Prepared by team; reviewed by Manager & EQCR"],
]
r = 4
for k, v in meta:
    a = ws.cell(row=r, column=1, value=k); a.font = Font(bold=True, color=NAVY)
    a.fill = fill(LIGHTBLUE); a.border = BORDER
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    b = ws.cell(row=r, column=2, value=v); b.border = BORDER
    b.alignment = Alignment(wrap_text=True, vertical="center", indent=1)
    r += 1

r += 1
idx_hdr = r
headers = ["Ref", "Sheet / Workpaper", "Purpose", "Std. Reference", "Preparer", "Status"]
rows = [
    ["WP-01", "01_Domain_Overview", "What digital audit & assurance covers; sub-domains & roles", "—", "Partner", "Final"],
    ["WP-02", "02_Engagement_Lifecycle", "End-to-end audit process, phases, deliverables & timing", "SA 200/300", "Manager", "Final"],
    ["WP-03", "03_Acceptance_Independence", "Client acceptance, independence & risk gate", "SA 210 / Code of Ethics", "Partner", "Final"],
    ["WP-04", "04_Risk_&_Materiality", "Risk assessment matrix + materiality computation", "SA 315 / SA 320", "Manager", "Reviewed"],
    ["WP-05", "05_ITGC_Testing", "IT General Controls testing (access, change, ops)", "SA 315 / ITGC", "IT Auditor", "In progress"],
    ["WP-06", "06_Application_Controls", "Automated / application control testing", "SA 330", "IT Auditor", "In progress"],
    ["WP-07", "07_JE_Data_Analytics", "Journal entry testing via data analytics (CAAT)", "SA 240 / SA 330", "Data Analyst", "In progress"],
    ["WP-08", "08_Substantive_Sampling", "Substantive test of details — revenue / AR sample", "SA 500 / SA 530", "Senior", "In progress"],
    ["WP-09", "09_Deliverables_Tracker", "Master tracker of engagement deliverables", "SA 230", "Manager", "Live"],
    ["WP-10", "10_Findings_Log", "Control deficiencies & audit findings register", "SA 265", "Manager", "Live"],
    ["WP-11", "11_KPIs_&_Insights", "Engagement KPIs, analytics & partner insights", "—", "Partner", "Live"],
    ["WP-12", "12_Tools_&_Tech", "Digital audit tool stack & emerging tech", "—", "Partner", "Final"],
    ["WP-13", "13_ESG_Assurance", "Limited assurance over sustainability/ESG metrics", "ISAE 3000 / IFRS S1-S2", "ESG Lead", "In progress"],
    ["WP-14", "14_SOC1_RCM", "Service org risk & control matrix (SOC 1 Type 2)", "ISAE 3402 / SSAE 18", "IT Auditor", "In progress"],
    ["WP-15", "15_Cyber_Maturity", "Cyber security maturity assessment (NIST CSF)", "NIST CSF 2.0", "Cyber Lead", "In progress"],
    ["WP-16", "16_Audit_Report", "Illustrative independent auditor's report", "SA 700/701/705/706", "Partner", "Draft"],
]
r = put_table(ws, idx_hdr, headers, rows,
              widths=[9, 26, 52, 22, 14, 13])
add_filter_freeze(ws, idx_hdr, 6)

# =========================================================================
# 01 - DOMAIN OVERVIEW
# =========================================================================
ws = wb.create_sheet("01_Domain_Overview")
title_block(ws, "DIGITAL AUDIT & ASSURANCE — DOMAIN OVERVIEW",
            "How technology has reshaped the assurance profile for a Chartered Accountant", 4)
headers = ["Sub-Domain", "What it means", "Typical CA activities", "Key deliverables"]
rows = [
    ["Financial Statement Audit (tech-enabled)",
     "Traditional statutory audit executed on a digital audit platform with 100% data ingestion instead of sampling where possible.",
     "Plan, risk-assess, test controls & balances, form opinion; use analytics on full ledgers.",
     "Audit opinion, financials, management letter"],
    ["IT Audit / ITGC & Application Controls",
     "Assurance over the IT environment supporting financial reporting (ERP, sub-ledgers, interfaces).",
     "Scope IT systems, test ITGCs (access/change/ops), automated controls, IPE reliability.",
     "ITGC report, control matrix, deficiency log"],
    ["Data Analytics & CAATs",
     "Using full-population analysis, scripts and visual analytics to identify risk and anomalies.",
     "JE testing, three-way match, duplicate/round-sum tests, trend & ratio analytics.",
     "Analytics workpapers, dashboards, exception lists"],
    ["SOC / SOX / Internal Controls Assurance",
     "Reporting on controls at service organisations (SOC 1/2) and internal control over financial reporting (ICFR/SOX).",
     "Design & operating effectiveness testing, walkthroughs, RCM build.",
     "SOC report, ICFR/IFC opinion, RCM"],
    ["Risk Advisory & Process Assurance",
     "Assurance over business processes, third parties, and regulatory compliance.",
     "Process walkthroughs, control gap analysis, remediation advisory.",
     "Risk & control matrix, remediation roadmap"],
    ["Emerging Assurance (ESG, AI, Cyber, Cloud)",
     "New assurance frontiers — sustainability/ESG reporting, AI governance, cyber & cloud posture.",
     "ESG data assurance, cyber maturity review, cloud config review.",
     "ESG assurance statement, cyber/cloud reports"],
]
nr = put_table(ws, 4, headers, rows, widths=[30, 44, 46, 34])
add_filter_freeze(ws, 4, 4)

# Skills block
nr += 1
ws.merge_cells(start_row=nr, start_column=1, end_row=nr, end_column=4)
c = ws.cell(row=nr, column=1, value="CORE SKILL-SET FOR A DIGITAL AUDIT CA")
c.font = hfont(12, WHITE, True); c.fill = fill(TEAL)
c.alignment = Alignment(indent=1, vertical="center"); ws.row_dimensions[nr].height = 22
nr += 1
skills = [
    ["Technical", "Ind AS/IFRS, SAs/ISAs, ICFR, revenue & consolidation, tax interplay"],
    ["IT & Data", "ERP (SAP/Oracle/D365), SQL, Excel/Power Query, Alteryx, Power BI/Tableau"],
    ["Analytics", "JE testing logic, full-population testing, anomaly & fraud analytics"],
    ["Governance", "ITGC frameworks (COBIT), SOC/SOX, cyber & cloud fundamentals"],
    ["Soft skills", "Professional scepticism, client handling, documentation discipline, review"],
]
nr = put_table(ws, nr, ["Area", "Detail"], skills, widths=[16, 100], header_fill=TEAL)

# =========================================================================
# 02 - ENGAGEMENT LIFECYCLE (END TO END)
# =========================================================================
ws = wb.create_sheet("02_Engagement_Lifecycle")
title_block(ws, "END-TO-END AUDIT ENGAGEMENT LIFECYCLE",
            "Phase-wise activities, standards, deliverables and indicative timing", 6)
headers = ["Phase", "Key activities", "Std. reference", "Primary deliverable", "Owner", "Indicative timing"]
rows = [
    ["1. Pre-engagement / Acceptance",
     "Client & engagement acceptance, independence checks, engagement letter, scoping, fee & team.",
     "SA 210, Ethics", "Signed engagement letter, acceptance memo", "Partner", "Wk 0-1"],
    ["2. Planning & Risk Assessment",
     "Understand entity & IT environment, risk assessment, materiality, audit strategy & plan, analytics scoping.",
     "SA 300/315/320", "Audit strategy & plan, materiality memo", "Manager", "Wk 1-3"],
    ["3. Controls Understanding & Testing",
     "Walkthroughs, RCM, ITGC & application control testing, evaluate design & operating effectiveness.",
     "SA 315/330", "RCM, ITGC report, control test WPs", "IT + Senior", "Wk 3-6"],
    ["4. Substantive Procedures",
     "Test of details & substantive analytics, JE testing, confirmations, estimates & provisions.",
     "SA 330/500/505/520/540", "Substantive WPs, exception logs", "Senior/Team", "Wk 5-9"],
    ["5. Completion & Review",
     "Subsequent events, going concern, management representations, EQCR, clearing review notes.",
     "SA 450/560/570/580/220", "Completion memo, review notes cleared", "Partner/EQCR", "Wk 9-11"],
    ["6. Reporting",
     "Form opinion, KAM, draft & issue audit report, communicate with TCWG, management letter.",
     "SA 700/701/705/706/260/265", "Audit report, KAM, management letter", "Partner", "Wk 11-12"],
    ["7. Post-issuance / Archival",
     "File assembly & archival, documentation retention, lessons learned, hot/cold file review.",
     "SA 230, SQM 1/2", "Archived audit file, quality review", "Manager", "Wk 12-13"],
]
nr = put_table(ws, 4, headers, rows, widths=[26, 52, 22, 34, 14, 14])
add_filter_freeze(ws, 4, 6)

# =========================================================================
# 03 - ACCEPTANCE & INDEPENDENCE
# =========================================================================
ws = wb.create_sheet("03_Acceptance_Independence")
title_block(ws, "CLIENT ACCEPTANCE & INDEPENDENCE — RISK GATE",
            "SA 210 / Code of Ethics — completed before engagement letter is signed", 5)
headers = ["#", "Consideration / Question", "Response", "Risk rating", "Comment"]
rows = [
    [1, "Integrity of management & TCWG assessed?", "Yes", "Low", "No adverse regulatory history noted."],
    [2, "Firm & team independence confirmed (financial, employment, business)?", "Yes", "Low", "Independence declarations obtained from all staff."],
    [3, "Non-audit services within permitted limits & fee cap?", "Yes", "Medium", "Tax compliance service pre-cleared; within cap."],
    [4, "Competence, capabilities & time available to perform?", "Yes", "Low", "IT audit + analytics specialists rostered."],
    [5, "Any conflicts of interest identified?", "No", "Low", "Conflict check run in firm system — clear."],
    [6, "Money-laundering / KYC checks completed?", "Yes", "Low", "Beneficial ownership verified."],
    [7, "Predecessor auditor communication done (if applicable)?", "N/A", "Low", "Continuing engagement — not applicable."],
    [8, "Fee reasonable vs. effort (no low-balling risk)?", "Yes", "Medium", "Recovery aligned with planned hours."],
    [9, "Group / component auditor arrangements clear?", "Yes", "Medium", "Two overseas components — instructions issued."],
    [10, "Overall: accept / decline / accept with conditions?", "Accept", "Medium", "Accept — enhanced review of revenue & IT."],
]
nr = put_table(ws, 4, headers, rows, widths=[5, 58, 14, 14, 46])
# conditional colour on risk
for row in ws.iter_rows(min_row=5, max_row=ws.max_row, min_col=4, max_col=4):
    for cell in row:
        v = str(cell.value).lower()
        if v == "low":   cell.fill = fill(GREEN_F); cell.font = Font(color=GREEN_T, bold=True)
        elif v == "medium": cell.fill = fill(AMBER_F); cell.font = Font(color=AMBER_T, bold=True)
        elif v == "high": cell.fill = fill(RED_F); cell.font = Font(color=RED_T, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
add_filter_freeze(ws, 4, 5)

# =========================================================================
# 04 - RISK & MATERIALITY
# =========================================================================
ws = wb.create_sheet("04_Risk_&_Materiality")
title_block(ws, "RISK ASSESSMENT MATRIX & MATERIALITY",
            "SA 315 (risk) + SA 320 (materiality) — figures illustrative (INR crore)", 6)

# Materiality computation block
ws.cell(row=4, column=1, value="A. MATERIALITY COMPUTATION").font = hfont(12, NAVY, True)
matrows = [
    ["Benchmark", "Revenue (PBT loss-making yr avoided)"],
    ["Benchmark amount (INR cr)", 4820],
    ["% applied", "0.5%"],
    ["Overall materiality (OM) (INR cr)", 24.1],
    ["Performance materiality (75% of OM) (INR cr)", 18.08],
    ["Clearly trivial threshold (5% of OM) (INR cr)", 1.21],
]
rr = 5
for k, v in matrows:
    a = ws.cell(row=rr, column=1, value=k); a.font = Font(bold=True, color=NAVY); a.fill = fill(LIGHTBLUE); a.border = BORDER
    b = ws.cell(row=rr, column=2, value=v); b.border = BORDER
    rr += 1
ws.column_dimensions["A"].width = 42; ws.column_dimensions["B"].width = 30

rr += 1
ws.cell(row=rr, column=1, value="B. SIGNIFICANT RISK ASSESSMENT MATRIX").font = hfont(12, NAVY, True)
rr += 1
headers = ["Risk area", "Assertion", "Inherent risk", "Control risk", "RoMM", "Planned response"]
rows = [
    ["Revenue recognition (cut-off, channel mix)", "Occurrence / Cut-off", "High", "Medium", "Significant", "Full-pop cut-off analytics + TOD on sample"],
    ["Inventory existence & valuation", "Existence / Valuation", "High", "Medium", "Significant", "Physical count attendance + NRV testing"],
    ["Management override of controls", "All", "High", "High", "Significant", "JE data analytics (mandatory per SA 240)"],
    ["Expected credit loss on receivables", "Valuation", "High", "Medium", "Significant", "Re-perform ECL model, challenge assumptions"],
    ["IT-dependent controls (ERP)", "All", "Medium", "Medium", "Elevated", "ITGC + application control testing"],
    ["Capitalisation of intangibles", "Existence / Cut-off", "Medium", "Low", "Normal", "Vouch additions, useful-life review"],
]
rr = put_table(ws, rr, headers, rows, widths=[38, 20, 14, 14, 15, 40])
# colour ratings
for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=3, max_col=5):
    for cell in row:
        if cell.value in ("High", "Significant"): cell.fill = fill(RED_F); cell.font = Font(color=RED_T, bold=True)
        elif cell.value in ("Medium", "Elevated"): cell.fill = fill(AMBER_F); cell.font = Font(color=AMBER_T, bold=True)
        elif cell.value in ("Low", "Normal"): cell.fill = fill(GREEN_F); cell.font = Font(color=GREEN_T, bold=True)

# =========================================================================
# 05 - ITGC TESTING
# =========================================================================
ws = wb.create_sheet("05_ITGC_Testing")
title_block(ws, "IT GENERAL CONTROLS (ITGC) TESTING WORKPAPER",
            "System in scope: SAP S/4HANA (financial reporting) | Sample-based operating effectiveness testing", 9)
headers = ["Ctrl ID", "ITGC Domain", "Control description", "Test procedure", "Population", "Sample", "Exceptions", "Result", "Deficiency severity"]
rows = [
    ["AC-01", "Access - Provisioning", "New user access approved by manager before grant.", "Inspect approval for sample of new joiners.", 142, 25, 0, "Pass", "None"],
    ["AC-02", "Access - Terminations", "Access revoked within 48 hrs of exit.", "Compare HR exit list to AD disable dates.", 63, 25, 3, "Fail", "Significant"],
    ["AC-03", "Access - Priv. Users", "Privileged/SoD access reviewed quarterly.", "Inspect 4 quarterly review sign-offs.", 4, 4, 1, "Fail", "Deficiency"],
    ["AC-04", "Access - Recertification", "Periodic user access recertification performed.", "Inspect half-yearly recertification evidence.", 2, 2, 0, "Pass", "None"],
    ["CM-01", "Change Mgmt - Approval", "Changes approved & tested before production.", "Inspect change tickets for approval & UAT.", 210, 30, 2, "Fail", "Deficiency"],
    ["CM-02", "Change Mgmt - Seg.", "Developers cannot migrate to production.", "Inspect migration access list vs dev list.", 18, 18, 0, "Pass", "None"],
    ["OP-01", "Ops - Batch Jobs", "Failed batch jobs monitored & resolved.", "Inspect job monitoring log for sample days.", 365, 25, 0, "Pass", "None"],
    ["OP-02", "Ops - Backups", "Daily backups taken & restore tested.", "Inspect backup logs + restore test report.", 365, 25, 0, "Pass", "None"],
    ["OP-03", "Ops - Incident Mgmt", "IT incidents logged, prioritised & closed.", "Inspect incident tickets & SLA adherence.", 480, 30, 4, "Fail", "Deficiency"],
]
nr = put_table(ws, 4, headers, rows, widths=[9, 20, 40, 40, 12, 10, 12, 10, 18])
for row in ws.iter_rows(min_row=5, max_row=ws.max_row, min_col=8, max_col=9):
    for cell in row:
        v = str(cell.value)
        if v in ("Pass", "None"): cell.fill = fill(GREEN_F); cell.font = Font(color=GREEN_T, bold=True)
        elif v in ("Fail", "Significant"): cell.fill = fill(RED_F); cell.font = Font(color=RED_T, bold=True)
        elif v == "Deficiency": cell.fill = fill(AMBER_F); cell.font = Font(color=AMBER_T, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
add_filter_freeze(ws, 4, 9)

# =========================================================================
# 06 - APPLICATION CONTROLS
# =========================================================================
ws = wb.create_sheet("06_Application_Controls")
title_block(ws, "AUTOMATED / APPLICATION CONTROL TESTING",
            "Configurable & automated controls within SAP — test of one (with ITGC reliance)", 8)
headers = ["Ctrl ID", "Process", "Automated control", "Config tested", "Test approach", "Expected", "Actual", "Result"]
rows = [
    ["APP-01", "Order-to-Cash", "3-way match (PO-GRN-Invoice) before payment", "Tolerance = 0%", "Reprocess 2 valid + 2 exception scenarios", "Block on mismatch", "Blocked correctly", "Pass"],
    ["APP-02", "Order-to-Cash", "Credit limit block on over-limit sales orders", "Limit per customer master", "Enter order above limit", "Order blocked", "Order blocked", "Pass"],
    ["APP-03", "Procure-to-Pay", "Duplicate invoice check on vendor+invoice no.", "Duplicate flag = ON", "Post duplicate invoice", "System warning/block", "Warning only, not block", "Deficiency"],
    ["APP-04", "Record-to-Report", "System auto-posts FX revaluation at month-end", "Rate source = central table", "Recompute for 1 period", "Matches system", "Matches system", "Pass"],
    ["APP-05", "Order-to-Cash", "Sales price derived from valid price master only", "Manual override = restricted", "Attempt manual price override", "Override blocked", "Override blocked", "Pass"],
    ["APP-06", "Payroll", "Net pay recomputation control on master change", "Recalc = automatic", "Change grade, verify recalputation", "Auto recalculated", "Auto recalculated", "Pass"],
]
nr = put_table(ws, 4, headers, rows, widths=[9, 18, 40, 22, 34, 22, 22, 12])
for row in ws.iter_rows(min_row=5, max_row=ws.max_row, min_col=8, max_col=8):
    for cell in row:
        v = str(cell.value)
        if v == "Pass": cell.fill = fill(GREEN_F); cell.font = Font(color=GREEN_T, bold=True)
        elif v == "Deficiency": cell.fill = fill(AMBER_F); cell.font = Font(color=AMBER_T, bold=True)
        elif v == "Fail": cell.fill = fill(RED_F); cell.font = Font(color=RED_T, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
add_filter_freeze(ws, 4, 8)

# =========================================================================
# 07 - JE DATA ANALYTICS
# =========================================================================
ws = wb.create_sheet("07_JE_Data_Analytics")
title_block(ws, "JOURNAL ENTRY TESTING — DATA ANALYTICS (CAAT)",
            "Full-population JE analysis (SA 240 mandatory) — sample of flagged entries below", 9)
headers = ["JE No.", "Post Date", "Posted By", "Amount (INR)", "Account", "Risk flag", "Reason flagged", "Follow-up status", "Conclusion"]
rows = [
    ["JE-100234", "31-Mar-2026", "svc_batch", 1250000, "Revenue", "Period-end", "Posted on last day, manual, to revenue", "Vouched", "Valid — genuine sale"],
    ["JE-100489", "31-Mar-2026", "r.mehta", 4890000, "Revenue", "Round-sum", "Exact round amount, manual, weekend post", "Queried", "Cut-off error — adjust"],
    ["JE-100512", "02-Apr-2026", "a.khan", 320000, "Provisions", "Back-dated", "Posted after year-end to Mar period", "Vouched", "Valid — accrual"],
    ["JE-100777", "15-Feb-2026", "admin_fi", 9999999, "Suspense", "Rare user", "Admin ID rarely posts, high value", "Escalated", "Reclass required"],
    ["JE-100801", "31-Mar-2026", "s.rao", 275500, "COGS", "Unusual pair", "Debit COGS / credit revenue combination", "Queried", "Investigate — pending"],
    ["JE-100950", "29-Mar-2026", "svc_batch", 60000, "Cash", "Weekend", "Posted on Sunday, manual", "Vouched", "Valid — bank sweep"],
    ["JE-101003", "31-Mar-2026", "r.mehta", 15000000, "Revenue", "Seldom-used acct", "Rarely used revenue GL, period-end", "Escalated", "Potential override"],
    ["JE-101120", "30-Mar-2026", "p.nair", 812000, "Expenses", "Reversal", "Posted then reversed next day", "Vouched", "Valid — correction"],
]
nr = put_table(ws, 4, headers, rows, widths=[12, 12, 12, 16, 12, 14, 40, 16, 30])
# number format on amount
for row in ws.iter_rows(min_row=5, max_row=ws.max_row, min_col=4, max_col=4):
    for cell in row:
        cell.number_format = '#,##0'
        cell.alignment = Alignment(horizontal="right")
# highlight escalated / adjust conclusions
for row in ws.iter_rows(min_row=5, max_row=ws.max_row, min_col=8, max_col=9):
    for cell in row:
        v = str(cell.value).lower()
        if "escalated" in v or "adjust" in v or "reclass" in v or "override" in v or "pending" in v:
            cell.fill = fill(RED_F); cell.font = Font(color=RED_T, bold=True)
        elif "valid" in v or "vouched" in v:
            cell.fill = fill(GREEN_F); cell.font = Font(color=GREEN_T, bold=True)
        elif "queried" in v:
            cell.fill = fill(AMBER_F); cell.font = Font(color=AMBER_T, bold=True)
add_filter_freeze(ws, 4, 9)

# analytics summary
nr += 1
ws.cell(row=nr, column=1, value="POPULATION SUMMARY").font = hfont(11, NAVY, True)
nr += 1
summ = [
    ["Total JEs in population", "418,930"],
    ["Total manual JEs", "37,214"],
    ["JEs meeting >=1 risk criterion", "1,046"],
    ["JEs selected for follow-up", "62"],
    ["Exceptions requiring adjustment", "4"],
    ["Net proposed adjustment (INR cr)", "6.3"],
]
put_table(ws, nr, ["Metric", "Value"], summ, widths=[36, 20], header_fill=TEAL)

# =========================================================================
# 08 - SUBSTANTIVE SAMPLING
# =========================================================================
ws = wb.create_sheet("08_Substantive_Sampling")
title_block(ws, "SUBSTANTIVE TEST OF DETAILS — REVENUE / ACCOUNTS RECEIVABLE",
            "SA 500/530 — MUS sample of revenue invoices with tie-out to source documents", 10)
headers = ["Sample #", "Invoice No.", "Customer", "Date", "Amount (INR)", "PO?", "GRN/POD?", "Recorded period OK?", "Diff (INR)", "Result"]
rows = [
    [1, "INV-55012", "MetroMart", "12-Mar-2026", 1820000, "Y", "Y", "Y", 0, "No exception"],
    [2, "INV-55098", "QuickShop", "28-Mar-2026", 2450000, "Y", "Y", "N", 2450000, "Cut-off error"],
    [3, "INV-55140", "UrbanGro", "05-Mar-2026", 990000, "Y", "Y", "Y", 0, "No exception"],
    [4, "INV-55201", "BigBasketX", "19-Mar-2026", 3120000, "Y", "Y", "Y", 0, "No exception"],
    [5, "INV-55233", "MetroMart", "31-Mar-2026", 4100000, "Y", "N", "N", 4100000, "No POD — cut-off"],
    [6, "INV-55290", "FreshCo", "22-Mar-2026", 760000, "Y", "Y", "Y", 0, "No exception"],
    [7, "INV-55345", "QuickShop", "08-Mar-2026", 1330000, "Y", "Y", "Y", 0, "No exception"],
    [8, "INV-55401", "UrbanGro", "27-Mar-2026", 2050000, "Y", "Y", "Y", 0, "No exception"],
    [9, "INV-55467", "PrimeRetail", "30-Mar-2026", 5600000, "N", "Y", "Y", 5600000, "No PO — verify"],
    [10, "INV-55510", "FreshCo", "15-Mar-2026", 1180000, "Y", "Y", "Y", 0, "No exception"],
]
nr = put_table(ws, 4, headers, rows, widths=[9, 12, 14, 12, 15, 6, 9, 16, 14, 16])
for row in ws.iter_rows(min_row=5, max_row=ws.max_row, min_col=5, max_col=5):
    for cell in row: cell.number_format = '#,##0'; cell.alignment = Alignment(horizontal="right")
for row in ws.iter_rows(min_row=5, max_row=ws.max_row, min_col=9, max_col=9):
    for cell in row: cell.number_format = '#,##0'; cell.alignment = Alignment(horizontal="right")
for row in ws.iter_rows(min_row=5, max_row=ws.max_row, min_col=10, max_col=10):
    for cell in row:
        v = str(cell.value)
        if v == "No exception": cell.fill = fill(GREEN_F); cell.font = Font(color=GREEN_T, bold=True)
        else: cell.fill = fill(RED_F); cell.font = Font(color=RED_T, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
add_filter_freeze(ws, 4, 10)

# =========================================================================
# 09 - DELIVERABLES TRACKER
# =========================================================================
ws = wb.create_sheet("09_Deliverables_Tracker")
title_block(ws, "ENGAGEMENT DELIVERABLES — MASTER TRACKER",
            "Live tracker of all client-facing & internal deliverables", 7)
headers = ["Deliverable", "Type", "Owner", "Due date", "Status", "% complete", "Notes"]
rows = [
    ["Engagement letter", "Client", "Partner", "10-Apr-2026", "Complete", "100%", "Signed & filed"],
    ["Audit strategy & plan", "Internal", "Manager", "25-Apr-2026", "Complete", "100%", "Reviewed by partner"],
    ["Materiality memo", "Internal", "Manager", "25-Apr-2026", "Complete", "100%", "OM = 24.1 cr"],
    ["Risk & control matrix (RCM)", "Internal", "Senior", "10-May-2026", "In progress", "80%", "IT controls pending"],
    ["ITGC audit report", "Client", "IT Auditor", "20-May-2026", "In progress", "70%", "3 deficiencies noted"],
    ["JE analytics workpaper", "Internal", "Data Analyst", "25-May-2026", "In progress", "65%", "4 exceptions to clear"],
    ["Substantive testing file", "Internal", "Senior", "05-Jun-2026", "In progress", "55%", "Cut-off issues found"],
    ["Management representation letter", "Client", "Partner", "20-Jun-2026", "Not started", "0%", "Draft after fieldwork"],
    ["Management letter (deficiencies)", "Client", "Manager", "22-Jun-2026", "Not started", "0%", "Compile from findings log"],
    ["Audit report & KAM", "Client", "Partner", "28-Jun-2026", "Not started", "0%", "Revenue likely KAM"],
    ["EQCR sign-off", "Internal", "EQCR", "27-Jun-2026", "Not started", "0%", "Pre-issuance"],
    ["File archival", "Internal", "Manager", "26-Jul-2026", "Not started", "0%", "Within 60 days"],
]
nr = put_table(ws, 4, headers, rows, widths=[34, 12, 14, 14, 14, 12, 34])
for row in ws.iter_rows(min_row=5, max_row=ws.max_row, min_col=5, max_col=5):
    for cell in row:
        v = str(cell.value)
        if v == "Complete": cell.fill = fill(GREEN_F); cell.font = Font(color=GREEN_T, bold=True)
        elif v == "In progress": cell.fill = fill(AMBER_F); cell.font = Font(color=AMBER_T, bold=True)
        else: cell.fill = fill(LIGHTGREY); cell.font = Font(color="808080", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
add_filter_freeze(ws, 4, 7)

# =========================================================================
# 10 - FINDINGS LOG
# =========================================================================
ws = wb.create_sheet("10_Findings_Log")
title_block(ws, "CONTROL DEFICIENCIES & AUDIT FINDINGS REGISTER",
            "SA 265 — communicated to TCWG in the management letter", 8)
headers = ["Ref", "Finding", "Root cause", "Impact", "Severity", "Recommendation", "Mgmt response", "Status"]
rows = [
    ["F-01", "Access not revoked within SLA for 3 leavers", "Manual, un-monitored deprovisioning", "Unauthorised access risk", "Significant", "Automate HR-IT deprovisioning workflow", "Agreed — Q2 FY27", "Open"],
    ["F-02", "Quarterly privileged access review incomplete", "No formal reviewer accountability", "SoD conflict risk", "Deficiency", "Assign owner; tool-based review", "Agreed", "Open"],
    ["F-03", "2 changes migrated without UAT sign-off", "Emergency change process misused", "Unauthorised change risk", "Deficiency", "Enforce emergency change back-approval", "Partially agreed", "Open"],
    ["F-04", "Duplicate invoice control warns but not blocks", "Config set to warning only", "Duplicate payment risk", "Deficiency", "Reconfigure to hard block", "Agreed — immediate", "In progress"],
    ["F-05", "Revenue cut-off errors at period-end", "Weak goods-dispatch cut-off procedure", "Revenue overstatement", "Significant", "Strengthen dispatch cut-off & review", "Agreed", "Adjustment posted"],
    ["F-06", "IT incidents breaching SLA not escalated", "No escalation matrix in tool", "Prolonged control downtime", "Deficiency", "Configure SLA escalation alerts", "Agreed", "Open"],
]
nr = put_table(ws, 4, headers, rows, widths=[7, 36, 30, 26, 15, 34, 20, 16])
for row in ws.iter_rows(min_row=5, max_row=ws.max_row, min_col=5, max_col=5):
    for cell in row:
        v = str(cell.value)
        if v == "Significant": cell.fill = fill(RED_F); cell.font = Font(color=RED_T, bold=True)
        elif v == "Deficiency": cell.fill = fill(AMBER_F); cell.font = Font(color=AMBER_T, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
add_filter_freeze(ws, 4, 8)

# =========================================================================
# 11 - KPIs & INSIGHTS
# =========================================================================
ws = wb.create_sheet("11_KPIs_&_Insights")
title_block(ws, "ENGAGEMENT KPIs & PARTNER INSIGHTS",
            "Quality, efficiency & value metrics + partner-level commentary", 4)
headers = ["KPI", "Target", "Actual", "Commentary"]
rows = [
    ["Budget-to-actual hours", "<=100%", "104%", "Overrun driven by ITGC deficiencies re-testing."],
    ["% audit via data analytics", ">=60%", "72%", "Full-pop JE + revenue analytics increased coverage."],
    ["Review notes cleared before EQCR", "100%", "88%", "12% open — substantive file to close."],
    ["Client PBC on-time %", ">=90%", "76%", "IT evidence delayed; escalate to CFO."],
    ["Deficiencies remediated in-year", ">=50%", "33%", "F-04 & F-05 progressing; rest next year."],
    ["Restatements / errors post-issue", "0", "0", "No post-issuance errors — strong quality."],
    ["Staff utilisation vs plan", "85-95%", "91%", "Healthy; specialists well-deployed."],
    ["Client satisfaction (survey)", ">=4.5/5", "4.6/5", "Valued analytics-led insights."],
]
nr = put_table(ws, 4, headers, rows, widths=[38, 14, 14, 62])

nr += 1
ws.merge_cells(start_row=nr, start_column=1, end_row=nr, end_column=4)
c = ws.cell(row=nr, column=1, value="PARTNER'S KEY INSIGHTS (34-YR PERSPECTIVE)")
c.font = hfont(12, WHITE, True); c.fill = fill(TEAL); c.alignment = Alignment(indent=1, vertical="center")
ws.row_dimensions[nr].height = 22
insights = [
    "Analytics is now the default, not the exception — plan the audit around 100% data ingestion; sampling is the fallback where data quality is poor.",
    "IT is inseparable from financial audit — if you cannot rely on ITGCs, control reliance collapses and substantive effort (and cost) balloons. Scope IT early.",
    "Professional scepticism scales with data — surfacing 1,000 flagged JEs is easy; the skill is designing precise risk criteria so the team investigates the right 60.",
    "Documentation quality = defence quality — 'if it isn't documented, it wasn't done.' The file must let a reviewer re-perform your judgement.",
    "The deliverable clients value most is insight, not the opinion — the management letter and process observations differentiate a Big 4 audit.",
    "Cut-off and revenue recognition remain the perennial risk — automation reduces error but shifts risk to configuration and estimates (ECL, provisions).",
    "Independence & acceptance is the cheapest risk to manage and the most expensive to get wrong — never compromise the front-end gate.",
    "Emerging assurance (ESG, AI governance, cyber) is the growth frontier for a CA — build these capabilities now.",
]
nr += 1
for i, ins in enumerate(insights):
    ws.merge_cells(start_row=nr, start_column=1, end_row=nr, end_column=4)
    cell = ws.cell(row=nr, column=1, value=f"{i+1}.  {ins}")
    cell.alignment = Alignment(wrap_text=True, vertical="top", indent=1)
    cell.border = BORDER
    if i % 2 == 1: cell.fill = fill(LIGHTGREY)
    ws.row_dimensions[nr].height = 30
    nr += 1

# =========================================================================
# 12 - TOOLS & TECH
# =========================================================================
ws = wb.create_sheet("12_Tools_&_Tech")
title_block(ws, "DIGITAL AUDIT TOOL STACK & EMERGING TECH",
            "Illustrative technology used across the engagement", 4)
headers = ["Category", "Representative tools", "Use in audit", "Skill priority for CA"]
rows = [
    ["Audit platform", "KPMG Clara, PwC Aura, Deloitte Omnia, EY Canvas", "Workflow, documentation, risk linkage", "High"],
    ["Data prep / ETL", "Alteryx, Power Query, Python (pandas)", "Ingest & cleanse client ledgers", "High"],
    ["Analytics / CAAT", "IDEA, ACL/Galvanize, SQL, Excel", "JE testing, duplicates, three-way match", "High"],
    ["Visualisation", "Power BI, Tableau, Spotfire", "Dashboards, trend & ratio analytics", "Medium"],
    ["ITGC / GRC", "SAP GRC, ServiceNow, AuditBoard", "Access, change, SoD, control tracking", "Medium"],
    ["Confirmations", "Confirmation.com, blockchain pilots", "Bank & AR external confirmations", "Medium"],
    ["Emerging", "GenAI copilots, process mining, ESG platforms", "Doc review, anomaly detection, ESG data", "Growing"],
]
nr = put_table(ws, 4, headers, rows, widths=[20, 44, 46, 20])
add_filter_freeze(ws, 4, 4)

# =========================================================================
# 13 - ESG ASSURANCE
# =========================================================================
ws = wb.create_sheet("13_ESG_Assurance")
title_block(ws, "ESG / SUSTAINABILITY ASSURANCE WORKPAPER",
            "Limited assurance under ISAE 3000; metrics aligned to IFRS S1/S2 & GRI — figures illustrative", 8)
headers = ["Metric", "Category", "Unit", "Reported value", "Source / system", "Procedure", "Deviation", "Conclusion"]
rows = [
    ["Scope 1 GHG emissions", "Environment", "tCO2e", 48250, "Fuel & fleet logs", "Recompute using DEFRA factors", "1.8%", "Within tolerance"],
    ["Scope 2 GHG (location-based)", "Environment", "tCO2e", 71200, "Utility bills", "Vouch to invoices + grid factor", "0.6%", "Within tolerance"],
    ["Scope 3 (purchased goods)", "Environment", "tCO2e", 512000, "Spend-based model", "Reperform model on sample", "9.4%", "Qualify — estimate uncertainty"],
    ["Total energy consumed", "Environment", "MWh", 190400, "BMS + bills", "Reconcile meter to bills", "0.9%", "Within tolerance"],
    ["Water withdrawal", "Environment", "kL", 356000, "Meter logs", "Sample meter reconciliation", "2.1%", "Within tolerance"],
    ["Lost-time injury frequency", "Social", "per mn hrs", 1.4, "EHS system", "Recalc from incident register", "0.0%", "Agreed"],
    ["% women in workforce", "Social", "%", 38.0, "HRIS", "Tie to headcount extract", "0.0%", "Agreed"],
    ["Board independence", "Governance", "%", 55.0, "Board records", "Inspect board composition", "0.0%", "Agreed"],
    ["Renewable energy share", "Environment", "%", 27.5, "PPA + RECs", "Vouch RECs to certificates", "3.2%", "Qualify — REC evidence gap"],
]
nr = put_table(ws, 4, headers, rows, widths=[26, 14, 10, 14, 18, 34, 12, 30])
for row in ws.iter_rows(min_row=5, max_row=ws.max_row, min_col=4, max_col=4):
    for cell in row: cell.number_format = '#,##0.0'; cell.alignment = Alignment(horizontal="right")
for row in ws.iter_rows(min_row=5, max_row=ws.max_row, min_col=8, max_col=8):
    for cell in row:
        v = str(cell.value).lower()
        if "qualify" in v: cell.fill = fill(AMBER_F); cell.font = Font(color=AMBER_T, bold=True)
        else: cell.fill = fill(GREEN_F); cell.font = Font(color=GREEN_T, bold=True)
add_filter_freeze(ws, 4, 8)
nr += 1
ws.merge_cells(start_row=nr, start_column=1, end_row=nr, end_column=8)
c = ws.cell(row=nr, column=1, value="Assurance conclusion: Limited assurance — nothing came to our attention except Scope 3 and renewable share, "
            "where estimation uncertainty and REC evidence gaps warrant a qualified conclusion. Recommend strengthening Scope 3 data lineage.")
c.font = Font(italic=True, color=NAVY); c.fill = fill(AMBER); c.alignment = Alignment(wrap_text=True, vertical="top", indent=1)
ws.row_dimensions[nr].height = 44

# =========================================================================
# 14 - SOC 1 RISK & CONTROL MATRIX
# =========================================================================
ws = wb.create_sheet("14_SOC1_RCM")
title_block(ws, "SOC 1 (ISAE 3402 / SSAE 18) RISK & CONTROL MATRIX",
            "Service organisation: cloud payroll processor | Type 2 — design + operating effectiveness", 9)
headers = ["Ctrl ID", "Control objective", "Control activity", "Frequency", "Test of design", "Test of op. effectiveness", "Sample", "Exceptions", "Result"]
rows = [
    ["SOC-01", "Only authorised users access the system", "RBAC + MFA enforced; access approved", "Continuous", "Design adequate", "Inspect access grants", 25, 0, "No exception"],
    ["SOC-02", "Changes are authorised & tested", "Change advisory board approval + UAT", "Per change", "Design adequate", "Inspect change tickets", 25, 1, "Exception"],
    ["SOC-03", "Data processed completely & accurately", "Input validation + batch totals reconciled", "Per run", "Design adequate", "Reperform batch reconciliation", 30, 0, "No exception"],
    ["SOC-04", "Backups ensure recoverability", "Daily backup + quarterly restore test", "Daily/Qtly", "Design adequate", "Inspect logs + restore test", 12, 0, "No exception"],
    ["SOC-05", "Incidents managed & communicated", "Incident tickets with SLA & client comms", "Per event", "Design adequate", "Inspect incident sample", 20, 0, "No exception"],
    ["SOC-06", "Data segregated between clients", "Logical tenant isolation in multi-tenant DB", "Continuous", "Design adequate", "Inspect isolation config", 5, 0, "No exception"],
    ["SOC-07", "Sub-service orgs monitored", "Annual review of sub-processor SOC reports", "Annual", "Design adequate", "Inspect review evidence", 1, 0, "No exception"],
]
nr = put_table(ws, 4, headers, rows, widths=[9, 34, 40, 12, 16, 30, 10, 12, 15])
for row in ws.iter_rows(min_row=5, max_row=ws.max_row, min_col=9, max_col=9):
    for cell in row:
        v = str(cell.value)
        if v == "No exception": cell.fill = fill(GREEN_F); cell.font = Font(color=GREEN_T, bold=True)
        else: cell.fill = fill(AMBER_F); cell.font = Font(color=AMBER_T, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
add_filter_freeze(ws, 4, 9)

# =========================================================================
# 15 - CYBER MATURITY
# =========================================================================
ws = wb.create_sheet("15_Cyber_Maturity")
title_block(ws, "CYBER SECURITY MATURITY ASSESSMENT (NIST CSF 2.0)",
            "Maturity scale 1 (Initial) - 5 (Optimised) | Current vs. target with gap", 6)
headers = ["CSF Function", "Sample control area", "Current", "Target", "Gap", "Priority action"]
rows = [
    ["Govern", "Cyber risk governance & roles", 3, 4, 1, "Formalise board-level cyber oversight"],
    ["Identify", "Asset inventory & data classification", 2, 4, 2, "Deploy CMDB; classify crown-jewel data"],
    ["Protect", "Identity, access & MFA", 3, 4, 1, "Extend MFA to all remote & privileged access"],
    ["Protect", "Vulnerability & patch management", 2, 4, 2, "Establish 30-day critical patch SLA"],
    ["Detect", "Security monitoring / SIEM", 3, 4, 1, "Tune SIEM use-cases; 24x7 SOC coverage"],
    ["Respond", "Incident response plan & drills", 2, 4, 2, "Run tabletop exercises twice a year"],
    ["Recover", "Backup, DR & resilience testing", 3, 4, 1, "Test full DR failover annually"],
]
nr = put_table(ws, 4, headers, rows, widths=[14, 34, 12, 12, 10, 46])
for row in ws.iter_rows(min_row=5, max_row=ws.max_row, min_col=5, max_col=5):
    for cell in row:
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if cell.value and cell.value >= 2: cell.fill = fill(RED_F); cell.font = Font(color=RED_T, bold=True)
        else: cell.fill = fill(AMBER_F); cell.font = Font(color=AMBER_T, bold=True)
for row in ws.iter_rows(min_row=5, max_row=ws.max_row, min_col=3, max_col=4):
    for cell in row: cell.alignment = Alignment(horizontal="center", vertical="center")
add_filter_freeze(ws, 4, 6)
nr += 1
ws.merge_cells(start_row=nr, start_column=1, end_row=nr, end_column=6)
c = ws.cell(row=nr, column=1, value="Overall maturity: 2.6 / 5.0 (Developing). Highest exposure: asset visibility, patching and IR readiness. "
            "Recommend a 12-month uplift roadmap prioritising Identify & Respond.")
c.font = Font(italic=True, color=NAVY); c.fill = fill(AMBER); c.alignment = Alignment(wrap_text=True, vertical="top", indent=1)
ws.row_dimensions[nr].height = 40

# =========================================================================
# 16 - AUDIT REPORT TEMPLATE
# =========================================================================
ws = wb.create_sheet("16_Audit_Report")
title_block(ws, "INDEPENDENT AUDITOR'S REPORT — ILLUSTRATIVE TEMPLATE",
            "Unmodified opinion with a Key Audit Matter (SA 700/701) — for illustration only", 2)
ws.column_dimensions["A"].width = 26
ws.column_dimensions["B"].width = 104
report = [
    ("Addressee", "To the Members of Nimbus Retail Ltd (illustrative)"),
    ("Report on the audit of the financial statements", ""),
    ("Opinion",
     "We have audited the financial statements of Nimbus Retail Ltd, which comprise the balance sheet as at 31 March 2026, "
     "the statement of profit and loss, cash flows and changes in equity, and notes. In our opinion the financial statements "
     "give a true and fair view in conformity with Ind AS of the state of affairs, profit, and cash flows for the year then ended."),
    ("Basis for opinion",
     "We conducted our audit in accordance with the Standards on Auditing (SAs). Our responsibilities are described in the "
     "Auditor's Responsibilities section. We are independent of the Company per the Code of Ethics and have fulfilled our "
     "ethical responsibilities. We believe the audit evidence obtained is sufficient and appropriate."),
    ("Key Audit Matter — Revenue recognition & cut-off",
     "Revenue is a significant risk due to high transaction volume across channels and cut-off pressure at year-end. "
     "Our response: full-population data analytics over the revenue ledger and journal entries, targeted tests of details on "
     "cut-off around period-end, and evaluation of the automated 3-way match control. Cut-off exceptions identified were "
     "adjusted by management."),
    ("Management's & TCWG responsibilities",
     "Management is responsible for the preparation of the financial statements, internal control, and assessing going concern. "
     "Those charged with governance oversee the financial reporting process."),
    ("Auditor's responsibilities",
     "Our objectives are to obtain reasonable assurance about whether the financial statements are free from material "
     "misstatement, whether due to fraud or error, and to issue a report with our opinion. We exercise professional judgement "
     "and maintain professional scepticism throughout."),
    ("Report on other legal & regulatory requirements",
     "As required by the Companies Act 2013 and CARO, we report on the matters specified therein (illustrative placeholder)."),
    ("Signature block",
     "For & on behalf of [Firm Name], Chartered Accountants | Firm Reg. No. XXXXX | [Partner Name], Partner | Membership No. XXXXXX | "
     "UDIN: XXXXXXXXXXXX | Place: [City] | Date: 28 June 2026"),
]
rr = 4
for label, body in report:
    a = ws.cell(row=rr, column=1, value=label)
    a.font = Font(bold=True, color=WHITE); a.fill = fill(BLUE)
    a.alignment = Alignment(wrap_text=True, vertical="top", indent=1); a.border = BORDER
    b = ws.cell(row=rr, column=2, value=body)
    b.alignment = Alignment(wrap_text=True, vertical="top", indent=1); b.border = BORDER
    if body:
        lines = max(1, len(body) // 95 + 1)
        ws.row_dimensions[rr].height = max(20, 15 * lines + 6)
    rr += 1

# Global: default row height & gridlines off for cleaner look on data sheets
for sheet in wb.worksheets:
    sheet.sheet_view.showGridLines = False

out = "/projects/sandbox/Digital_Audit_Assurance_Deliverables_Big4.xlsx"
wb.save(out)
print("Saved:", out)
print("Sheets:", [s.title for s in wb.worksheets])
